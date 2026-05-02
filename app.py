import calendar
import pickle
from pathlib import Path
from datetime import date, datetime
from typing import Optional

import pandas as pd
import streamlit as st

import io
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

st.set_page_config(page_title="Presenze e Pagamenti Automatizzati", layout="wide")

# =========================
# COSTANTI
# =========================

LOGO_URL = "https://raw.githubusercontent.com/GlucaTekmar/operativita-pdv/refs/heads/main/logo.png"

COLOR_ROSSO = "#C62828"
BG_APP = "#EEF1F4"
BG_BOX = "#F7F8FA"
BG_FIELD = "#FFFFFF"
BORDER_COLOR = "#97A3AF"
BORDER_COLOR_STRONG = "#7F8C99"
TEXT_COLOR = "#111111"
TEXT_MUTED = "#425466"

GIORNI_SETTIMANA = ["LUN", "MAR", "MER", "GIO", "VEN", "SAB", "DOM"]
MESI = {
    1: "Gennaio",
    2: "Febbraio",
    3: "Marzo",
    4: "Aprile",
    5: "Maggio",
    6: "Giugno",
    7: "Luglio",
    8: "Agosto",
    9: "Settembre",
    10: "Ottobre",
    11: "Novembre",
    12: "Dicembre",
}

TIPI_ASSENZA = [
    "",
    "Ferie",
    "Permesso",
    "Malattia",
    "Infortunio",
    "L. 104",
    "Festivo non lavorato",
    "Assenza",
]

COLONNE_MASTER_EDICOLA = [
    "LUN", "MAR", "MER", "GIO", "VEN", "SAB", "DOM",
    "COMPENSO_MENSILE", "NETTO_ORA", "ID_GDMS", "AGENZIA", "PDV",
    "MHS_TITOLARE", "TEL_MHS", "MAIL_MHS", "COD_FISCALE",
    "TIPO_CONTRATTO", "SCADENZA_CONTRATTO",
]

COLONNE_MASTER_LIBRI = [
    "ID_GDMS", "PDV", "MHS_TITOLARE", "TEL_MHS", "MAIL_MHS",
    "COD_FISCALE", "AGENZIA", "TIPO_LIBRI", "ELEMENTI_BANCO",
    "LUN", "MAR", "MER", "GIO", "VEN", "SAB", "DOM", "NETTO_ORA",
    "TIPO_CONTRATTO", "SCADENZA_CONTRATTO",
]

COLONNE_MASTER_SPOT = [
    "MHS_SOSTITUTO_SPOT", "COD_FISCALE", "TEL", "MAIL", "NETTO_ORA",
    "TIPO_CONTRATTO", "SCADENZA_CONTRATTO",
]

HEADER_ALIASES = {
    "CODICE_FISCALE": "COD_FISCALE",
    "CF": "COD_FISCALE",
    "MHS TITOLARE": "MHS_TITOLARE",
    "TEL MHS": "TEL_MHS",
    "MAIL MHS": "MAIL_MHS",
    "TIPO CONTRATTO": "TIPO_CONTRATTO",
    "SCADENZA CONTRATTO": "SCADENZA_CONTRATTO",
    "NETTO ORA": "NETTO_ORA",
    "COMPENSO MENSILE": "COMPENSO_MENSILE",
    "TIPO LIBRI": "TIPO_LIBRI",
    "ELEMENTI BANCO": "ELEMENTI_BANCO",
}

ORIGINE_EDICOLA = "EDICOLA"
ORIGINE_LIBRI = "LIBRI"

COLONNE_TABELLA_TECNICHE = [
    "GIORNO_NUM",
    "DATA",
    "GIORNO_SETTIMANA",
    "DATA_VIEW",
    "EDICOLA_ORE",
    "EDICOLA_€",
    "EDICOLA_TIPO_ASSENZA",
    "MONDADORI_ORE",
    "MONDADORI_€",
    "MONDADORI_TIPO_ASSENZA",
    "GIUNTI_ORE",
    "GIUNTI_€",
    "GIUNTI_TIPO_ASSENZA",
    "FESTIVO",
    "ROW_STATUS",
    "STEP4_SELECTED_DAY",
    "BASE_EDICOLA_ORE",
    "BASE_MONDADORI_ORE",
    "BASE_GIUNTI_ORE",
]

# =========================
# STILE
# =========================

def inject_global_css():
    st.markdown(
        f"""
        <style>
            :root {{
                --app-bg: {BG_APP};
                --box-bg: {BG_BOX};
                --field-bg: {BG_FIELD};
                --border: {BORDER_COLOR};
                --border-strong: {BORDER_COLOR_STRONG};
                --text: {TEXT_COLOR};
                --muted: {TEXT_MUTED};
                --red: {COLOR_ROSSO};
            }}

            html, body, [class*="css"], .stApp {{
                color: var(--text);
                font-size: 16px !important;
            }}

            .stApp {{
                background: var(--app-bg);
            }}

            .block-container {{
                padding-top: 0.8rem;
                padding-bottom: 2rem;
                max-width: 96%;
            }}

            section[data-testid="stSidebar"] {{
                background: #E3E8EE !important;
                border-right: 2px solid var(--border-strong);
            }}

            section[data-testid="stSidebar"] * {{
                font-size: 18px !important;
                color: var(--text) !important;
            }}

            .sidebar-logo-wrap {{
                padding-top: 6px;
                padding-bottom: 12px;
                border-bottom: 2px solid var(--border-strong);
                margin-bottom: 14px;
            }}

            .main-header-box {{
                background: var(--box-bg);
                border: 2px solid var(--border-strong);
                border-radius: 16px;
                padding: 16px 18px;
                margin-bottom: 18px;
            }}

            .page-title-box {{
                background: var(--box-bg);
                border: 2px solid var(--border-strong);
                border-radius: 16px;
                padding: 14px 16px;
                margin-bottom: 16px;
            }}

            .main-header-title {{
                font-size: 2rem;
                font-weight: 800;
                color: var(--red);
                margin: 0;
                line-height: 1.1;
            }}

            .main-header-subtitle {{
                font-size: 1.25rem;
                font-weight: 700;
                color: var(--text);
                margin: 0.35rem 0 0 0;
            }}

            .section-title {{
                font-size: 1.65rem;
                font-weight: 800;
                color: var(--red);
                margin: 0;
                line-height: 1.15;
            }}

            .section-box {{
                background: var(--box-bg);
                border: 2px solid var(--border-strong);
                border-radius: 16px;
                padding: 16px;
                margin-bottom: 16px;
            }}

            .inner-box {{
                background: #F4F6F8;
                border: 2px solid var(--border-strong);
                border-radius: 14px;
                padding: 14px;
                margin-bottom: 14px;
            }}

            .soft-note {{
                border-left: 4px solid var(--red);
                background: #FFF5F5;
                border-radius: 10px;
                padding: 12px 14px;
                font-size: 1rem;
                margin-bottom: 12px;
            }}

            .action-box {{
                background: #FFF4F4;
                border: 3px solid {COLOR_ROSSO};
                border-radius: 16px;
                padding: 16px;
                margin-bottom: 16px;
            }}

            .action-title {{
                font-size: 1.35rem;
                font-weight: 900;
                color: {COLOR_ROSSO};
                margin: 0 0 0.35rem 0;
            }}

            .action-subtitle {{
                font-size: 1rem;
                font-weight: 700;
                color: {TEXT_COLOR};
                margin: 0 0 0.8rem 0;
            }}

            .warning-box {{
                background: #FDECEC;
                border: 2px solid #D9534F;
                color: #8B0000;
                border-radius: 12px;
                padding: 12px 14px;
                font-size: 1rem;
                font-weight: 800;
                margin-bottom: 10px;
            }}

            .table-title {{
                font-size: 1.32rem;
                font-weight: 800;
                color: var(--text);
                margin-bottom: 0.6rem;
            }}

            div[data-baseweb="select"] > div,
            .stTextInput input,
            .stNumberInput input,
            .stTextArea textarea,
            [data-testid="stFileUploaderDropzone"],
            [data-testid="stDataFrame"],
            [data-testid="stTable"] {{
                background: var(--field-bg) !important;
                border: 2px solid var(--border-strong) !important;
                border-radius: 12px !important;
                color: var(--text) !important;
                -webkit-text-fill-color: var(--text) !important;
                opacity: 1 !important;
            }}

            .stTextInput input:disabled,
            .stNumberInput input:disabled,
            .stTextArea textarea:disabled,
            div[data-baseweb="select"] input:disabled {{
                color: var(--text) !important;
                -webkit-text-fill-color: var(--text) !important;
                opacity: 1 !important;
            }}

            .stTextInput label p,
            .stNumberInput label p,
            .stTextArea label p,
            .stFileUploader label p,
            .stSelectbox label p,
            .stCheckbox label p {{
                color: var(--text) !important;
                font-size: 1.06rem !important;
                font-weight: 800 !important;
            }}

            .stButton > button {{
                background: white !important;
                color: var(--text) !important;
                border: 2px solid var(--border-strong) !important;
                border-radius: 12px !important;
                font-size: 1.04rem !important;
                font-weight: 800 !important;
                min-height: 46px;
            }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_header():
    st.markdown('<div class="main-header-box">', unsafe_allow_html=True)
    col_logo, col_text = st.columns([1, 5])
    with col_logo:
        st.image(LOGO_URL, width=150)
    with col_text:
        st.markdown(
            """
            <p class="main-header-title">DASHBOARD ADMIN</p>
            <p class="main-header-subtitle">presenze e pagamenti automatizzati</p>
            """,
            unsafe_allow_html=True,
        )
    st.markdown("</div>", unsafe_allow_html=True)


def render_page_title(title: str):
    st.markdown(
        f"""
        <div class="page-title-box">
            <p class="section-title">{title}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_warning_box(message: str):
    st.markdown(
        f'<div class="warning-box">ATTENZIONE — {message}</div>',
        unsafe_allow_html=True,
    )

# =========================
# UTILITY
# =========================

def normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_upper(value) -> str:
    return normalize_text(value).upper()


def safe_float(value) -> float:
    if pd.isna(value) or str(value).strip() == "":
        return 0.0
    text = str(value).strip().replace("€", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def normalize_date_text(value) -> str:
    if pd.isna(value) or str(value).strip() == "":
        return ""
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    cleaned = []
    for col in df.columns:
        new_col = str(col).strip().replace("\n", " ")
        new_col = " ".join(new_col.split())
        new_col = HEADER_ALIASES.get(new_col.upper(), new_col)
        cleaned.append(new_col)
    df.columns = cleaned
    return df


def validate_columns(df: pd.DataFrame, required_columns: list[str]) -> list[str]:
    return [col for col in required_columns if col not in df.columns]


def get_day_label(day_date: date) -> str:
    return GIORNI_SETTIMANA[day_date.weekday()]


def get_festivo_multiplier(agenzia: str) -> float:
    agenzia_up = normalize_upper(agenzia)
    if agenzia_up == "TEKMAR":
        return 1.30
    if agenzia_up == "UP":
        return 1.20
    return 1.00


def calculate_row_amount(ore: float, netto_ora: float, festivo: bool, agenzia: str) -> float:
    base = round(safe_float(ore) * safe_float(netto_ora), 2)
    if festivo:
        return round(base * get_festivo_multiplier(agenzia), 2)
    return base


def format_sheet_key(row_id: str, anno: int, mese: int) -> str:
    return f"{row_id}__{anno}__{mese:02d}"


def format_scadenza(value: str) -> str:
    return normalize_text(value)


def get_selected_days(anno: int, mese: int, modo: str, giorno_singolo_value, periodo_value) -> list[int]:
    if modo == "Giorno singolo":
        if giorno_singolo_value is None:
            return []
        if giorno_singolo_value.year != anno or giorno_singolo_value.month != mese:
            return []
        return [int(giorno_singolo_value.day)]

    if not periodo_value or len(periodo_value) != 2:
        return []
    data_inizio, data_fine = periodo_value
    if not data_inizio or not data_fine or data_inizio > data_fine:
        return []

    primo_giorno = date(anno, mese, 1)
    ultimo_giorno = date(anno, mese, calendar.monthrange(anno, mese)[1])
    if data_inizio < primo_giorno or data_fine > ultimo_giorno:
        return []

    giorni = []
    current = data_inizio
    while current <= data_fine:
        giorni.append(current.day)
        current = date.fromordinal(current.toordinal() + 1)
    return giorni


def selected_days_text(selected_days: list[int]) -> str:
    if not selected_days:
        return ""
    if len(selected_days) == 1:
        return f"Giorno {selected_days[0]}"
    return f"Giorni da {min(selected_days)} a {max(selected_days)}"

# =========================
# NORMALIZZAZIONE MASTER
# =========================

def normalize_master_edicola(df: pd.DataFrame) -> pd.DataFrame:
    df = clean_columns(df).copy()
    for col in ["ID_GDMS", "AGENZIA", "PDV", "MHS_TITOLARE", "TEL_MHS", "MAIL_MHS", "COD_FISCALE", "TIPO_CONTRATTO"]:
        df[col] = df[col].apply(normalize_text)
    df["SCADENZA_CONTRATTO"] = df["SCADENZA_CONTRATTO"].apply(normalize_date_text)
    for col in GIORNI_SETTIMANA + ["COMPENSO_MENSILE", "NETTO_ORA"]:
        df[col] = df[col].apply(safe_float)
    return df


def normalize_master_libri(df: pd.DataFrame) -> pd.DataFrame:
    df = clean_columns(df).copy()
    for col in ["ID_GDMS", "PDV", "MHS_TITOLARE", "TEL_MHS", "MAIL_MHS", "COD_FISCALE", "AGENZIA", "TIPO_LIBRI", "TIPO_CONTRATTO"]:
        df[col] = df[col].apply(normalize_text)
    df["ELEMENTI_BANCO"] = df["ELEMENTI_BANCO"].apply(normalize_text)
    df["SCADENZA_CONTRATTO"] = df["SCADENZA_CONTRATTO"].apply(normalize_date_text)
    for col in GIORNI_SETTIMANA + ["NETTO_ORA"]:
        df[col] = df[col].apply(safe_float)
    df["TIPO_LIBRI"] = df["TIPO_LIBRI"].str.upper().str.strip()
    return df


def normalize_master_spot(df: pd.DataFrame) -> pd.DataFrame:
    df = clean_columns(df).copy()
    for col in ["MHS_SOSTITUTO_SPOT", "COD_FISCALE", "TEL", "MAIL", "TIPO_CONTRATTO"]:
        df[col] = df[col].apply(normalize_text)
    df["SCADENZA_CONTRATTO"] = df["SCADENZA_CONTRATTO"].apply(normalize_date_text)
    df["NETTO_ORA"] = df["NETTO_ORA"].apply(safe_float)
    return df

# =========================
# SESSIONE / STORAGE
# =========================

STORAGE_DIR = Path("/var/data") if Path("/var/data").exists() else Path(".storage")
STATE_FILE = STORAGE_DIR / "app_state.pkl"

PERSIST_KEYS = [
    "df_edicola",
    "df_libri",
    "df_spot",
    "generation_table",
    "fogli_generati",
    "foglio_attivo",
    "master_loaded",
    "sheet_warnings",
    "anno_lavoro",
    "mese_lavoro",
]

def save_app_state():
    STORAGE_DIR.mkdir(parents=True, exist_ok=True)
    payload = {key: st.session_state.get(key) for key in PERSIST_KEYS}
    temp_file = STATE_FILE.with_suffix(".tmp")
    with open(temp_file, "wb") as f:
        pickle.dump(payload, f)
    temp_file.replace(STATE_FILE)

def load_app_state():
    if not STATE_FILE.exists():
        return
    with open(STATE_FILE, "rb") as f:
        payload = pickle.load(f)
    if not isinstance(payload, dict):
        return
    for key in PERSIST_KEYS:
        if key in payload:
            st.session_state[key] = payload[key]

def get_mese_anno_lavoro() -> tuple[int, int]:
    return int(st.session_state["anno_lavoro"]), int(st.session_state["mese_lavoro"])

def ensure_session_state():
    defaults = {
        "df_edicola": pd.DataFrame(),
        "df_libri": pd.DataFrame(),
        "df_spot": pd.DataFrame(),
        "generation_table": pd.DataFrame(),
        "fogli_generati": {},
        "foglio_attivo": None,
        "master_loaded": False,
        "sheet_warnings": {},
        "anno_lavoro": 2026,
        "mese_lavoro": 1,
        "_storage_loaded_once": False,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

    if not st.session_state["_storage_loaded_once"]:
        load_app_state()
        st.session_state["_storage_loaded_once"] = True

# =========================
# GENERATION TABLE
# =========================

def build_generation_table(df_edicola: pd.DataFrame, df_libri: pd.DataFrame) -> pd.DataFrame:
    righe = []

    for idx, row in df_edicola.reset_index(drop=True).iterrows():
        righe.append(
            {
                "SELEZIONA": False,
                "ROW_ID": f"EDICOLA_{idx}",
                "ORIGINE": "Master",
                "ORIGINE_MASTER": ORIGINE_EDICOLA,
                "TIPO_LIBRI": "",
                "NOME": normalize_text(row["MHS_TITOLARE"]),
                "COD_FISCALE": normalize_text(row["COD_FISCALE"]),
                "SOCIETA": normalize_text(row["AGENZIA"]),
                "PDV": normalize_text(row["PDV"]),
                "TELEFONO": normalize_text(row["TEL_MHS"]),
                "EMAIL": normalize_text(row["MAIL_MHS"]),
                "NETTO_ORA": safe_float(row["NETTO_ORA"]),
                "NETTO_MESE": safe_float(row["COMPENSO_MENSILE"]),
                "TIPO_CONTRATTO": normalize_text(row["TIPO_CONTRATTO"]),
                "SCADENZA_CONTRATTO": format_scadenza(row["SCADENZA_CONTRATTO"]),
                "ATTIVITA_RIGA": "EDICOLA",
                "MASTER_INDEX": idx,
                "IS_STEP4": False,
            }
        )

    for idx, row in df_libri.reset_index(drop=True).iterrows():
        tipo_libri = normalize_upper(row["TIPO_LIBRI"])
        attivita_riga = "LIBRI"
        if tipo_libri in {"MONDADORI", "GIUNTI"}:
            attivita_riga = f"LIBRI - {tipo_libri}"

        righe.append(
            {
                "SELEZIONA": False,
                "ROW_ID": f"LIBRI_{idx}",
                "ORIGINE": "Master",
                "ORIGINE_MASTER": ORIGINE_LIBRI,
                "TIPO_LIBRI": normalize_text(row["TIPO_LIBRI"]),
                "NOME": normalize_text(row["MHS_TITOLARE"]),
                "COD_FISCALE": normalize_text(row["COD_FISCALE"]),
                "SOCIETA": normalize_text(row["AGENZIA"]),
                "PDV": normalize_text(row["PDV"]),
                "TELEFONO": normalize_text(row["TEL_MHS"]),
                "EMAIL": normalize_text(row["MAIL_MHS"]),
                "NETTO_ORA": safe_float(row["NETTO_ORA"]),
                "NETTO_MESE": 0.0,
                "TIPO_CONTRATTO": normalize_text(row["TIPO_CONTRATTO"]),
                "SCADENZA_CONTRATTO": format_scadenza(row["SCADENZA_CONTRATTO"]),
                "ATTIVITA_RIGA": attivita_riga,
                "MASTER_INDEX": idx,
                "IS_STEP4": False,
            }
        )

    result = pd.DataFrame(righe)
    if result.empty:
        return result
    return result.sort_values(
        ["ORIGINE_MASTER", "NOME", "SOCIETA", "PDV", "TIPO_LIBRI"]
    ).reset_index(drop=True)


def append_record_to_generation_table(record: dict):
    # I fogli creati da pagina 4 non devono più alimentare
    # nessuna tabella dedicata in pagina 2.
    return

# =========================
# COSTRUZIONE / CALCOLI FOGLI
# =========================

def build_presence_dataframe_from_master_row(
    row: pd.Series,
    df_edicola: pd.DataFrame,
    df_libri: pd.DataFrame,
    anno: int,
    mese: int,
) -> pd.DataFrame:
    giorni_mese = calendar.monthrange(anno, mese)[1]
    righe = []

    origine = normalize_upper(row["ORIGINE_MASTER"])
    tipo_libri = normalize_upper(row.get("TIPO_LIBRI", ""))
    societa = normalize_text(row["SOCIETA"])
    netto_ora = safe_float(row["NETTO_ORA"])
    master_index = int(row["MASTER_INDEX"])

    source_row = df_edicola.iloc[master_index] if origine == ORIGINE_EDICOLA else df_libri.iloc[master_index]

    def get_master_hours_for_day(giorno_label: str) -> tuple[float, float, float]:
        if origine == ORIGINE_EDICOLA:
            return safe_float(source_row[giorno_label]), 0.0, 0.0
        if tipo_libri == "MONDADORI":
            return 0.0, safe_float(source_row[giorno_label]), 0.0
        if tipo_libri == "GIUNTI":
            return 0.0, 0.0, safe_float(source_row[giorno_label])
        return 0.0, 0.0, 0.0

    for giorno_num in range(1, giorni_mese + 1):
        current_date = date(anno, mese, giorno_num)
        giorno_label = get_day_label(current_date)
        festivo_default = current_date.weekday() == 6

        edicola_ore, mondadori_ore, giunti_ore = get_master_hours_for_day(giorno_label)

        righe.append(
            {
                "GIORNO_NUM": giorno_num,
                "DATA": current_date.strftime("%d/%m/%Y"),
                "GIORNO_SETTIMANA": giorno_label,
                "DATA_VIEW": current_date.strftime("%d/%m/%Y"),
                "EDICOLA_ORE": edicola_ore,
                "EDICOLA_€": calculate_row_amount(edicola_ore, netto_ora, festivo_default, societa),
                "EDICOLA_TIPO_ASSENZA": "",
                "MONDADORI_ORE": mondadori_ore,
                "MONDADORI_€": calculate_row_amount(mondadori_ore, netto_ora, festivo_default, societa),
                "MONDADORI_TIPO_ASSENZA": "",
                "GIUNTI_ORE": giunti_ore,
                "GIUNTI_€": calculate_row_amount(giunti_ore, netto_ora, festivo_default, societa),
                "GIUNTI_TIPO_ASSENZA": "",
                "FESTIVO": festivo_default,
                "ROW_STATUS": "",
                "STEP4_SELECTED_DAY": False,
                "BASE_EDICOLA_ORE": edicola_ore,
                "BASE_MONDADORI_ORE": mondadori_ore,
                "BASE_GIUNTI_ORE": giunti_ore,
            }
        )
    return pd.DataFrame(righe, columns=COLONNE_TABELLA_TECNICHE)


def build_empty_presence_dataframe(anno: int, mese: int, selected_days: list[int]) -> pd.DataFrame:
    righe = []
    giorni_mese = calendar.monthrange(anno, mese)[1]
    for giorno_num in range(1, giorni_mese + 1):
        current_date = date(anno, mese, giorno_num)
        festivo_default = current_date.weekday() == 6
        righe.append(
            {
                "GIORNO_NUM": giorno_num,
                "DATA": current_date.strftime("%d/%m/%Y"),
                "GIORNO_SETTIMANA": get_day_label(current_date),
                "DATA_VIEW": current_date.strftime("%d/%m/%Y"),
                "EDICOLA_ORE": 0.0,
                "EDICOLA_€": 0.0,
                "EDICOLA_TIPO_ASSENZA": "",
                "MONDADORI_ORE": 0.0,
                "MONDADORI_€": 0.0,
                "MONDADORI_TIPO_ASSENZA": "",
                "GIUNTI_ORE": 0.0,
                "GIUNTI_€": 0.0,
                "GIUNTI_TIPO_ASSENZA": "",
                "FESTIVO": festivo_default,
                "ROW_STATUS": "",
                "STEP4_SELECTED_DAY": giorno_num in selected_days,
                "BASE_EDICOLA_ORE": 0.0,
                "BASE_MONDADORI_ORE": 0.0,
                "BASE_GIUNTI_ORE": 0.0,
            }
        )
    return pd.DataFrame(righe, columns=COLONNE_TABELLA_TECNICHE)


def calculate_sheet_stats(df: pd.DataFrame, origine_master: str, netto_ora: float = 0.0, societa: str = "") -> dict:
    if origine_master == ORIGINE_EDICOLA:
        current_hours = df["EDICOLA_ORE"].apply(safe_float)
        base_hours = df["BASE_EDICOLA_ORE"].apply(safe_float)

        current_euro_series = df["EDICOLA_€"].apply(safe_float)
        base_euro_series = df.apply(
            lambda row: calculate_row_amount(
                safe_float(row["BASE_EDICOLA_ORE"]),
                netto_ora,
                bool(row["FESTIVO"]),
                societa,
            ),
            axis=1,
        )
    else:
        current_hours = df["MONDADORI_ORE"].apply(safe_float) + df["GIUNTI_ORE"].apply(safe_float)
        base_hours = df["BASE_MONDADORI_ORE"].apply(safe_float) + df["BASE_GIUNTI_ORE"].apply(safe_float)

        current_euro_series = df["MONDADORI_€"].apply(safe_float) + df["GIUNTI_€"].apply(safe_float)
        base_euro_series = df.apply(
            lambda row: (
                calculate_row_amount(
                    safe_float(row["BASE_MONDADORI_ORE"]),
                    netto_ora,
                    bool(row["FESTIVO"]),
                    societa,
                )
                + calculate_row_amount(
                    safe_float(row["BASE_GIUNTI_ORE"]),
                    netto_ora,
                    bool(row["FESTIVO"]),
                    societa,
                )
            ),
            axis=1,
        )

    giorni_lavorati_attuali = int((current_hours > 0).sum())
    giorni_lavorati_base = int((base_hours > 0).sum())
    giorni_modificati = int((df["ROW_STATUS"].astype(str).str.strip() != "").sum())

    tot_ore_attuali = round(current_hours.sum(), 2)
    tot_ore_base = round(base_hours.sum(), 2)
    tot_ore_ridotte = round((base_hours - current_hours).clip(lower=0).sum(), 2)

    tot_euro_attuali = round(current_euro_series.sum(), 2)
    tot_euro_base = round(base_euro_series.sum(), 2)
    tot_euro_scalati = round((base_euro_series - current_euro_series).clip(lower=0).sum(), 2)

    return {
        "GIORNI_LAVORATI_ATTUALI": giorni_lavorati_attuali,
        "GIORNI_LAVORATI_BASE": giorni_lavorati_base,
        "GIORNI_MODIFICATI": giorni_modificati,
        "TOT_ORE_ATTUALI": tot_ore_attuali,
        "TOT_ORE_BASE": tot_ore_base,
        "TOT_ORE_RIDOTTE": tot_ore_ridotte,
        "TOT_ATTIVITA_ATTUALE_€": tot_euro_attuali,
        "TOT_ATTIVITA_BASE_€": tot_euro_base,
        "TOT_€_DA_SCALARE": tot_euro_scalati,
    }


def get_row_status(row: pd.Series, origine_master: str) -> str:
    has_step4 = bool(row.get("STEP4_SELECTED_DAY", False))
    has_red = False

    if origine_master == ORIGINE_EDICOLA:
        base = safe_float(row["BASE_EDICOLA_ORE"])
        current = safe_float(row["EDICOLA_ORE"])
        ass = normalize_text(row["EDICOLA_TIPO_ASSENZA"]) != ""
        if current < base or ass:
            has_red = True
    else:
        base = safe_float(row["BASE_MONDADORI_ORE"]) + safe_float(row["BASE_GIUNTI_ORE"])
        current = safe_float(row["MONDADORI_ORE"]) + safe_float(row["GIUNTI_ORE"])
        ass = normalize_text(row["MONDADORI_TIPO_ASSENZA"]) != "" or normalize_text(row["GIUNTI_TIPO_ASSENZA"]) != ""
        if current < base or ass:
            has_red = True

    if has_step4 and has_red:
        return "🟡🔴"
    if has_step4:
        return "🟡"
    if has_red:
        return "🔴"
    return ""


def apply_presence_rules(tabella: pd.DataFrame, netto_ora: float, societa: str, origine_master: str, allow_increase: bool) -> tuple[pd.DataFrame, list[str]]:
    df = tabella.copy()
    warnings = []

    for col in ["EDICOLA_ORE", "MONDADORI_ORE", "GIUNTI_ORE"]:
        df[col] = df[col].apply(safe_float)

    for idx in df.index:
        if origine_master == ORIGINE_EDICOLA:
            base = safe_float(df.at[idx, "BASE_EDICOLA_ORE"])
            current = safe_float(df.at[idx, "EDICOLA_ORE"])

            if not allow_increase and current > base:
                df.at[idx, "EDICOLA_ORE"] = base
                warnings.append(f"Giorno {int(df.at[idx, 'GIORNO_NUM'])}: aumento ore non consentito.")

            if safe_float(df.at[idx, "EDICOLA_ORE"]) > 0:
                df.at[idx, "EDICOLA_TIPO_ASSENZA"] = ""

            df.at[idx, "MONDADORI_ORE"] = 0.0
            df.at[idx, "MONDADORI_€"] = 0.0
            df.at[idx, "MONDADORI_TIPO_ASSENZA"] = ""
            df.at[idx, "GIUNTI_ORE"] = 0.0
            df.at[idx, "GIUNTI_€"] = 0.0
            df.at[idx, "GIUNTI_TIPO_ASSENZA"] = ""

            df.at[idx, "EDICOLA_€"] = calculate_row_amount(
                safe_float(df.at[idx, "EDICOLA_ORE"]),
                netto_ora,
                bool(df.at[idx, "FESTIVO"]),
                societa,
            )

        else:
            base_mon = safe_float(df.at[idx, "BASE_MONDADORI_ORE"])
            base_giu = safe_float(df.at[idx, "BASE_GIUNTI_ORE"])

            if not allow_increase and safe_float(df.at[idx, "MONDADORI_ORE"]) > base_mon:
                df.at[idx, "MONDADORI_ORE"] = base_mon
                warnings.append(f"Giorno {int(df.at[idx, 'GIORNO_NUM'])}: aumento ore non consentito.")

            if not allow_increase and safe_float(df.at[idx, "GIUNTI_ORE"]) > base_giu:
                df.at[idx, "GIUNTI_ORE"] = base_giu
                warnings.append(f"Giorno {int(df.at[idx, 'GIORNO_NUM'])}: aumento ore non consentito.")

            if safe_float(df.at[idx, "MONDADORI_ORE"]) > 0:
                df.at[idx, "MONDADORI_TIPO_ASSENZA"] = ""
            if safe_float(df.at[idx, "GIUNTI_ORE"]) > 0:
                df.at[idx, "GIUNTI_TIPO_ASSENZA"] = ""

            df.at[idx, "EDICOLA_ORE"] = 0.0
            df.at[idx, "EDICOLA_€"] = 0.0
            df.at[idx, "EDICOLA_TIPO_ASSENZA"] = ""

            df.at[idx, "MONDADORI_€"] = calculate_row_amount(
                safe_float(df.at[idx, "MONDADORI_ORE"]),
                netto_ora,
                bool(df.at[idx, "FESTIVO"]),
                societa,
            )
            df.at[idx, "GIUNTI_€"] = calculate_row_amount(
                safe_float(df.at[idx, "GIUNTI_ORE"]),
                netto_ora,
                bool(df.at[idx, "FESTIVO"]),
                societa,
            )

        df.at[idx, "ROW_STATUS"] = get_row_status(df.loc[idx], origine_master)
        base_data = normalize_text(df.at[idx, "DATA"])
        status = df.at[idx, "ROW_STATUS"]
        df.at[idx, "DATA_VIEW"] = f"{status} {base_data}" if status else base_data

    if warnings:
        warnings = list(dict.fromkeys(warnings))

    return df, warnings


def update_sheet_totals(record: dict):
    allow_increase = bool(record.get("allow_increase", False))
    origine_master = normalize_upper(record["origine_master"])

    # SOLO per foglio vuoto STEP 4:
    # quando inserisco nuove ore, quelle ore diventano la nuova BASE del foglio.
    # Questo permette:
    # - netto mese fisso dopo valorizzazione
    # - giorni lavorati base corretti
    # - ore ridotte/azzerate corrette
    # - € da scalare corretti
    if record.get("is_step4", False) and record.get("netto_mese_dynamic", False):
        df_base = record["tabella"].copy()

        if origine_master == ORIGINE_EDICOLA:
            for idx in df_base.index:
                current = safe_float(df_base.at[idx, "EDICOLA_ORE"])
                base = safe_float(df_base.at[idx, "BASE_EDICOLA_ORE"])
                if current > base:
                    df_base.at[idx, "BASE_EDICOLA_ORE"] = current
        else:
            for idx in df_base.index:
                current_mon = safe_float(df_base.at[idx, "MONDADORI_ORE"])
                base_mon = safe_float(df_base.at[idx, "BASE_MONDADORI_ORE"])
                if current_mon > base_mon:
                    df_base.at[idx, "BASE_MONDADORI_ORE"] = current_mon

                current_giu = safe_float(df_base.at[idx, "GIUNTI_ORE"])
                base_giu = safe_float(df_base.at[idx, "BASE_GIUNTI_ORE"])
                if current_giu > base_giu:
                    df_base.at[idx, "BASE_GIUNTI_ORE"] = current_giu

        record["tabella"] = df_base

    record["tabella"], warnings = apply_presence_rules(
        tabella=record["tabella"],
        netto_ora=record["netto_ora"],
        societa=record["societa"],
        origine_master=origine_master,
        allow_increase=allow_increase,
    )

    stats = calculate_sheet_stats(
        df=record["tabella"],
        origine_master=origine_master,
        netto_ora=record["netto_ora"],
        societa=record["societa"],
    )

    if record.get("is_step4", False) and record.get("netto_mese_dynamic", False):
        # FOGLIO VUOTO:
        # - giorni lavorati = base storica costruita inserendo le ore
        # - netto mese = base storica € e poi resta fisso
        record["giorni_lavorati"] = stats["GIORNI_LAVORATI_BASE"]
        record["netto_mese"] = round(safe_float(stats["TOT_ATTIVITA_BASE_€"]), 2)
    else:
        # FOGLI STANDARD / SOSTITUZIONI
        record["giorni_lavorati"] = stats["GIORNI_LAVORATI_ATTUALI"]

        # Per sostituzioni STEP 4:
        # appena creato il foglio, netto mese = totale corpo iniziale e poi resta fisso
        if record.get("is_step4", False) and safe_float(record.get("netto_mese", 0.0)) == 0 and safe_float(stats["TOT_ATTIVITA_BASE_€"]) > 0:
            record["netto_mese"] = round(safe_float(stats["TOT_ATTIVITA_BASE_€"]), 2)

    record["giorni_modificati"] = stats["GIORNI_MODIFICATI"]
    record["tot_ore_lavorative_mese"] = stats["TOT_ORE_ATTUALI"]
    record["tot_ore_azzerate"] = stats["TOT_ORE_RIDOTTE"]
    record["tot_euro_da_scalare"] = stats["TOT_€_DA_SCALARE"]
    record["tot_attivita"] = stats["TOT_ATTIVITA_ATTUALE_€"]

    # Totale finale = totale ATTUALE del corpo + fondo foglio
    record["tot_netto_mese"] = round(
        safe_float(record["tot_attivita"])
        + safe_float(record["arretrati"])
        + safe_float(record["extra"])
        + safe_float(record["affiancamenti"])
        + safe_float(record["domeniche"])
        + safe_float(record["rimborso"]),
        2,
    )

    return warnings


def clear_entire_sheet(record: dict):
    df = record["tabella"].copy()

    df["EDICOLA_ORE"] = 0.0
    df["EDICOLA_€"] = 0.0
    df["EDICOLA_TIPO_ASSENZA"] = ""

    df["MONDADORI_ORE"] = 0.0
    df["MONDADORI_€"] = 0.0
    df["MONDADORI_TIPO_ASSENZA"] = ""

    df["GIUNTI_ORE"] = 0.0
    df["GIUNTI_€"] = 0.0
    df["GIUNTI_TIPO_ASSENZA"] = ""

    record["tabella"] = df
    update_sheet_totals(record)

# =========================
# RECORD HELPERS
# =========================

def init_sheet_record_from_master(
    row: pd.Series,
    df_edicola: pd.DataFrame,
    df_libri: pd.DataFrame,
    anno: int,
    mese: int,
) -> dict:
    tabella = build_presence_dataframe_from_master_row(
        row=row,
        df_edicola=df_edicola,
        df_libri=df_libri,
        anno=anno,
        mese=mese,
    )

    record = {
        "row_id": normalize_text(row["ROW_ID"]),
        "origine_master": normalize_text(row["ORIGINE_MASTER"]),
        "tipo_libri": normalize_text(row["TIPO_LIBRI"]),
        "attivita_riga": normalize_text(row["ATTIVITA_RIGA"]),
        "societa": normalize_text(row["SOCIETA"]),
        "mese": mese,
        "anno": anno,
        "nome": normalize_text(row["NOME"]),
        "cf": normalize_text(row["COD_FISCALE"]),
        "pdv": normalize_text(row["PDV"]),
        "telefono": normalize_text(row["TELEFONO"]),
        "email": normalize_text(row["EMAIL"]),
        "netto_mese": round(safe_float(row["NETTO_MESE"]), 2),
        "netto_ora": round(safe_float(row["NETTO_ORA"]), 2),
        "tipo_contratto": normalize_text(row["TIPO_CONTRATTO"]),
        "scadenza_contratto": format_scadenza(row["SCADENZA_CONTRATTO"]),
        "giorni_lavorati": 0,
        "giorni_modificati": 0,
        "tot_ore_lavorative_mese": 0.0,
        "tot_ore_azzerate": 0.0,
        "tot_euro_da_scalare": 0.0,
        "lucchetto_foglio": False,
        "lucchetto_mese": False,
        "arretrati": 0.0,
        "extra": 0.0,
        "affiancamenti": 0.0,
        "domeniche": 0.0,
        "rimborso": 0.0,
        "rimborso_allegati": [],
        "note_generali": "",
        "tabella": tabella,
        "tot_attivita": 0.0,
        "tot_netto_mese": 0.0,
        "is_step4": False,
        "allow_increase": False,
        "netto_mese_dynamic": False,
    }
    update_sheet_totals(record)

    if normalize_upper(record["origine_master"]) == ORIGINE_LIBRI and safe_float(record["netto_mese"]) == 0:
        record["netto_mese"] = round(safe_float(record["tot_attivita"]), 2)

    return record


def base_step4_record(modalita: str, societa: str, attivita: str, anno: int, mese: int, selected_days: list[int]) -> dict:
    row_id = f"STEP4_{anno}_{mese:02d}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    is_empty = normalize_upper(modalita) == "FOGLIO PRESENZA VUOTO"
    return {
        "row_id": row_id,
        "origine_master": attivita,
        "tipo_libri": "",
        "attivita_riga": attivita,
        "societa": societa,
        "mese": mese,
        "anno": anno,
        "nome": "",
        "cf": "",
        "pdv": "",
        "telefono": "",
        "email": "",
        "netto_mese": 0.0,
        "netto_ora": 0.0,
        "tipo_contratto": "",
        "scadenza_contratto": "",
        "giorni_lavorati": 0,
        "giorni_modificati": 0,
        "tot_ore_lavorative_mese": 0.0,
        "tot_ore_azzerate": 0.0,
        "tot_euro_da_scalare": 0.0,
        "lucchetto_foglio": False,
        "lucchetto_mese": False,
        "arretrati": 0.0,
        "extra": 0.0,
        "affiancamenti": 0.0,
        "domeniche": 0.0,
        "rimborso": 0.0,
        "rimborso_allegati": [],
        "note_generali": "",
        "tabella": build_empty_presence_dataframe(anno, mese, selected_days),
        "tot_attivita": 0.0,
        "tot_netto_mese": 0.0,
        "is_step4": True,
        "step4_modalita": normalize_text(modalita),
        "step4_selected_days": list(selected_days),
        "allow_increase": is_empty,          # foglio vuoto compilabile
        "netto_mese_dynamic": is_empty,      # foglio vuoto: netto mese segue il corpo
    }

# =========================
# STEP 4 HELPERS
# =========================

def get_societa_options() -> list[str]:
    societa = set()
    if not st.session_state["df_edicola"].empty:
        societa.update(st.session_state["df_edicola"]["AGENZIA"].dropna().astype(str).str.strip().tolist())
    if not st.session_state["df_libri"].empty:
        societa.update(st.session_state["df_libri"]["AGENZIA"].dropna().astype(str).str.strip().tolist())
    return sorted([x for x in societa if normalize_text(x) != ""])


def build_dipendenti_table(societa: str, attivita: str, modalita: str) -> pd.DataFrame:
    righe = []
    societa_up = normalize_upper(societa)
    attivita_up = normalize_upper(attivita)
    modalita_up = normalize_upper(modalita)

    if modalita_up == "SOSTITUZIONE CON SOSTITUTO SPOT":
        df = st.session_state["df_spot"].copy()
        for idx, row in df.reset_index(drop=True).iterrows():
            righe.append(
                {
                    "SELEZIONA": False,
                    "SOURCE_ID": f"SPOT_{idx}",
                    "NOMINATIVO": normalize_text(row["MHS_SOSTITUTO_SPOT"]),
                    "CODICE_FISCALE": normalize_text(row["COD_FISCALE"]),
                    "TIPO_CONTRATTO": normalize_text(row["TIPO_CONTRATTO"]),
                    "SCADENZA_CONTRATTO": normalize_text(row["SCADENZA_CONTRATTO"]),
                    "NETTO_ORA": safe_float(row["NETTO_ORA"]),
                    "TEL": normalize_text(row["TEL"]),
                    "MAIL": normalize_text(row["MAIL"]),
                }
            )
    else:
        if attivita_up == ORIGINE_EDICOLA:
            df = st.session_state["df_edicola"].copy()
            df = df[df["AGENZIA"].astype(str).str.upper().str.strip() == societa_up]
        else:
            df = st.session_state["df_libri"].copy()
            df = df[df["AGENZIA"].astype(str).str.upper().str.strip() == societa_up]

        for idx, row in df.reset_index(drop=True).iterrows():
            righe.append(
                {
                    "SELEZIONA": False,
                    "SOURCE_ID": f"{attivita_up}_{idx}",
                    "NOMINATIVO": normalize_text(row["MHS_TITOLARE"]),
                    "CODICE_FISCALE": normalize_text(row["COD_FISCALE"]),
                    "TIPO_CONTRATTO": normalize_text(row["TIPO_CONTRATTO"]),
                    "SCADENZA_CONTRATTO": normalize_text(row["SCADENZA_CONTRATTO"]),
                    "NETTO_ORA": safe_float(row["NETTO_ORA"]),
                    "TEL": normalize_text(row["TEL_MHS"]),
                    "MAIL": normalize_text(row["MAIL_MHS"]),
                }
            )

    return pd.DataFrame(righe)


def build_pdv_table(societa: str, attivita: str) -> pd.DataFrame:
    righe = []
    societa_up = normalize_upper(societa)
    attivita_up = normalize_upper(attivita)

    if attivita_up == ORIGINE_EDICOLA:
        df = st.session_state["df_edicola"].copy()
        df = df[df["AGENZIA"].astype(str).str.upper().str.strip() == societa_up]
        for idx, row in df.reset_index(drop=True).iterrows():
            righe.append(
                {
                    "SELEZIONA": False,
                    "SOURCE_ID": f"PDV_EDICOLA_{idx}",
                    "PUNTO_VENDITA": normalize_text(row["PDV"]),
                    "TIPO_ATTIVITA": "EDICOLA",
                    "LUNEDI": safe_float(row["LUN"]),
                    "MARTEDI": safe_float(row["MAR"]),
                    "MERCOLEDI": safe_float(row["MER"]),
                    "GIOVEDI": safe_float(row["GIO"]),
                    "VENERDI": safe_float(row["VEN"]),
                    "SABATO": safe_float(row["SAB"]),
                    "DOMENICA": safe_float(row["DOM"]),
                    "MASTER_INDEX": idx,
                }
            )
    else:
        df = st.session_state["df_libri"].copy()
        df = df[df["AGENZIA"].astype(str).str.upper().str.strip() == societa_up]
        for idx, row in df.reset_index(drop=True).iterrows():
            righe.append(
                {
                    "SELEZIONA": False,
                    "SOURCE_ID": f"PDV_LIBRI_{idx}",
                    "PUNTO_VENDITA": normalize_text(row["PDV"]),
                    "TIPO_ATTIVITA": normalize_text(row["TIPO_LIBRI"]) or "LIBRI",
                    "LUNEDI": safe_float(row["LUN"]),
                    "MARTEDI": safe_float(row["MAR"]),
                    "MERCOLEDI": safe_float(row["MER"]),
                    "GIOVEDI": safe_float(row["GIO"]),
                    "VENERDI": safe_float(row["VEN"]),
                    "SABATO": safe_float(row["SAB"]),
                    "DOMENICA": safe_float(row["DOM"]),
                    "MASTER_INDEX": idx,
                    "TIPO_LIBRI": normalize_text(row["TIPO_LIBRI"]),
                }
            )
    return pd.DataFrame(righe)


def get_single_selected_row(df: pd.DataFrame) -> tuple[Optional[pd.Series], int]:
    if df.empty or "SELEZIONA" not in df.columns:
        return None, 0
    selected = df[df["SELEZIONA"] == True].copy()
    count = len(selected)
    if count == 1:
        return selected.iloc[0], 1
    return None, count


def apply_selected_rows_to_record(record: dict, dip_row: Optional[pd.Series], pdv_row: Optional[pd.Series], selected_days: list[int]):
    origine = normalize_upper(record["origine_master"])

    if dip_row is not None:
        record["nome"] = normalize_text(dip_row["NOMINATIVO"])
        record["cf"] = normalize_text(dip_row["CODICE_FISCALE"])
        record["tipo_contratto"] = normalize_text(dip_row["TIPO_CONTRATTO"])
        record["scadenza_contratto"] = normalize_text(dip_row["SCADENZA_CONTRATTO"])
        record["netto_ora"] = safe_float(dip_row["NETTO_ORA"])
        record["telefono"] = normalize_text(dip_row["TEL"])
        record["email"] = normalize_text(dip_row["MAIL"])

    if pdv_row is not None:
        record["pdv"] = normalize_text(pdv_row["PUNTO_VENDITA"])
        if origine == ORIGINE_LIBRI:
            record["tipo_libri"] = normalize_text(pdv_row.get("TIPO_LIBRI", ""))
            tipo = normalize_upper(record["tipo_libri"])
            if tipo in {"MONDADORI", "GIUNTI"}:
                record["attivita_riga"] = f"LIBRI - {tipo}"
            else:
                record["attivita_riga"] = "LIBRI"
        else:
            record["attivita_riga"] = "EDICOLA"

        weekday_map = {
            "LUN": "LUNEDI",
            "MAR": "MARTEDI",
            "MER": "MERCOLEDI",
            "GIO": "GIOVEDI",
            "VEN": "VENERDI",
            "SAB": "SABATO",
            "DOM": "DOMENICA",
        }

        for idx in record["tabella"].index:
            giorno_num = int(record["tabella"].at[idx, "GIORNO_NUM"])
            if giorno_num not in selected_days:
                continue
            giorno_dt = date(record["anno"], record["mese"], giorno_num)
            giorno_label = get_day_label(giorno_dt)
            ore = safe_float(pdv_row[weekday_map[giorno_label]])

            if origine == ORIGINE_EDICOLA:
                record["tabella"].at[idx, "EDICOLA_ORE"] = ore
                record["tabella"].at[idx, "BASE_EDICOLA_ORE"] = ore
            else:
                tipo = normalize_upper(record.get("tipo_libri", ""))
                if tipo == "MONDADORI":
                    record["tabella"].at[idx, "MONDADORI_ORE"] = ore
                    record["tabella"].at[idx, "BASE_MONDADORI_ORE"] = ore
                elif tipo == "GIUNTI":
                    record["tabella"].at[idx, "GIUNTI_ORE"] = ore
                    record["tabella"].at[idx, "BASE_GIUNTI_ORE"] = ore

# =========================
# HELPERS PAGINA 5 / CHIUSURA MESE
# =========================

def get_month_sheet_keys(anno: int, mese: int) -> list[str]:
    keys = []
    for foglio_key, record in st.session_state["fogli_generati"].items():
        if int(record.get("anno", 0)) == int(anno) and int(record.get("mese", 0)) == int(mese):
            keys.append(foglio_key)
    return keys


def get_record_origin_label(record: dict) -> str:
    return "Nuovo foglio/sostituzione" if bool(record.get("is_step4", False)) else "Master"


def get_record_attivita_export(record: dict) -> str:
    return normalize_text(record.get("attivita_riga", ""))


def get_record_day_hours(record: dict, day_num: int) -> float:
    df = record["tabella"]
    row = df[df["GIORNO_NUM"] == int(day_num)]
    if row.empty:
        return 0.0

    r = row.iloc[0]
    if normalize_upper(record.get("origine_master", "")) == ORIGINE_EDICOLA:
        return round(safe_float(r.get("EDICOLA_ORE", 0.0)), 2)

    return round(
        safe_float(r.get("MONDADORI_ORE", 0.0)) + safe_float(r.get("GIUNTI_ORE", 0.0)),
        2,
    )


def get_record_festivo_maggiorazione(record: dict) -> float:
    df = record["tabella"]
    netto_ora = safe_float(record.get("netto_ora", 0.0))
    totale = 0.0

    for _, row in df.iterrows():
        if not bool(row.get("FESTIVO", False)):
            continue

        if normalize_upper(record.get("origine_master", "")) == ORIGINE_EDICOLA:
            ore = safe_float(row.get("EDICOLA_ORE", 0.0))
            euro = safe_float(row.get("EDICOLA_€", 0.0))
            base = round(ore * netto_ora, 2)
            totale += max(round(euro - base, 2), 0.0)
        else:
            ore_mon = safe_float(row.get("MONDADORI_ORE", 0.0))
            euro_mon = safe_float(row.get("MONDADORI_€", 0.0))
            base_mon = round(ore_mon * netto_ora, 2)

            ore_giu = safe_float(row.get("GIUNTI_ORE", 0.0))
            euro_giu = safe_float(row.get("GIUNTI_€", 0.0))
            base_giu = round(ore_giu * netto_ora, 2)

            totale += max(round(euro_mon - base_mon, 2), 0.0)
            totale += max(round(euro_giu - base_giu, 2), 0.0)

    return round(totale, 2)


def get_record_netto_mese_export(record: dict) -> float:
    # REGOLA DEFINITIVA DEL FILE EXCEL UTENTE:
    # - origine Master -> campo netto_mese del foglio
    # - origine Nuovo foglio/sostituzione -> totale attività
    if get_record_origin_label(record) == "Master":
        return round(safe_float(record.get("netto_mese", 0.0)), 2)
    return round(safe_float(record.get("tot_attivita", 0.0)), 2)


def build_chiusura_mese_table_df(anno: int, mese: int) -> pd.DataFrame:
    righe = []

    for foglio_key in get_month_sheet_keys(anno, mese):
        record = st.session_state["fogli_generati"][foglio_key]
        righe.append(
            {
                "SOCIETA'": normalize_text(record.get("societa", "")),
                "ORIGINE": get_record_origin_label(record),
                "ATTIVITA'": get_record_attivita_export(record),
                "PDV": normalize_text(record.get("pdv", "")),
                "NOME": normalize_text(record.get("nome", "")),
                "TELEFONO": normalize_text(record.get("telefono", "")),
                "EMAIL": normalize_text(record.get("email", "")),
                "CF": normalize_text(record.get("cf", "")),
            }
        )

    df = pd.DataFrame(righe)
    if df.empty:
        return df

    return df.sort_values(["SOCIETA'", "ATTIVITA'", "PDV", "NOME"]).reset_index(drop=True)


def build_chiusura_mese_export_rows(anno: int, mese: int) -> list[dict]:
    righe = []
    giorni_mese = calendar.monthrange(anno, mese)[1]

    for foglio_key in get_month_sheet_keys(anno, mese):
        record = st.session_state["fogli_generati"][foglio_key]

        row = {
            "SOCIETA'": normalize_text(record.get("societa", "")),
            "ORIGINE": get_record_origin_label(record),
            "ATTIVITA'": get_record_attivita_export(record),
            "TIPO CONTRATTO": normalize_text(record.get("tipo_contratto", "")),
            "SCADENZA CONTRATTO": normalize_text(record.get("scadenza_contratto", "")),
            "NOMINATIVO": normalize_text(record.get("nome", "")),
            "COD. FISCALE": normalize_text(record.get("cf", "")),
            "TELEFONO": normalize_text(record.get("telefono", "")),
            "EMAIL": normalize_text(record.get("email", "")),
        }

        for day_num in range(1, giorni_mese + 1):
            row[date(anno, mese, day_num)] = round(get_record_day_hours(record, day_num), 2)

        row["NETTO ORARIO"] = round(safe_float(record.get("netto_ora", 0.0)), 2)
        row["NETTO MESE"] = get_record_netto_mese_export(record)
        row["MAGGIORAZIONE FESTIVO"] = get_record_festivo_maggiorazione(record)
        row["ASSENZE"] = round(safe_float(record.get("tot_euro_da_scalare", 0.0)), 2)
        row["ARRETRATO"] = round(safe_float(record.get("arretrati", 0.0)), 2)
        row["EXTRA"] = round(safe_float(record.get("extra", 0.0)), 2)
        row["AFFIANCAMENTO"] = round(safe_float(record.get("affiancamenti", 0.0)), 2)
        row["DOMENICHE"] = round(safe_float(record.get("domeniche", 0.0)), 2)
        row["RIMBORSI"] = round(safe_float(record.get("rimborso", 0.0)), 2)
        row["TOT. MESE DA PAGARE"] = None
        row["NOTE DEL MESE"] = normalize_text(record.get("note_generali", ""))

        righe.append(row)

    return righe


def build_chiusura_mese_xlsx_bytes(anno: int, mese: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{MESI[mese]} {anno}"

    giorni_mese = calendar.monthrange(anno, mese)[1]

    headers = [
        "SOCIETA'",
        "ORIGINE",
        "ATTIVITA'",
        "TIPO CONTRATTO",
        "SCADENZA CONTRATTO",
        "NOMINATIVO",
        "COD. FISCALE",
        "TELEFONO",
        "EMAIL",
    ]
    headers.extend([date(anno, mese, day_num) for day_num in range(1, giorni_mese + 1)])
    headers.extend(
        [
            "NETTO ORARIO",
            "NETTO MESE",
            "MAGGIORAZIONE FESTIVO",
            "ASSENZE",
            "ARRETRATO",
            "EXTRA",
            "AFFIANCAMENTO",
            "DOMENICHE",
            "RIMBORSI",
            "TOT. MESE DA PAGARE",
            "NOTE DEL MESE",
        ]
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        if isinstance(header, date):
            cell.number_format = "dd/mm/yyyy"

    rows = build_chiusura_mese_export_rows(anno, mese)

    # Mappa colonne per formula finale
    header_positions = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    col_netto_mese = header_positions["NETTO MESE"]
    col_magg = header_positions["MAGGIORAZIONE FESTIVO"]
    col_assenze = header_positions["ASSENZE"]
    col_arretrato = header_positions["ARRETRATO"]
    col_extra = header_positions["EXTRA"]
    col_aff = header_positions["AFFIANCAMENTO"]
    col_dom = header_positions["DOMENICHE"]
    col_rimb = header_positions["RIMBORSI"]
    col_tot_paga = header_positions["TOT. MESE DA PAGARE"]

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header, 0.0 if isinstance(header, date) else "")
            ws.cell(row=row_idx, column=col_idx, value=value)

        formula = (
            f"={get_column_letter(col_netto_mese)}{row_idx}"
            f"+{get_column_letter(col_magg)}{row_idx}"
            f"-{get_column_letter(col_assenze)}{row_idx}"
            f"+{get_column_letter(col_arretrato)}{row_idx}"
            f"+{get_column_letter(col_extra)}{row_idx}"
            f"+{get_column_letter(col_aff)}{row_idx}"
            f"+{get_column_letter(col_dom)}{row_idx}"
            f"+{get_column_letter(col_rimb)}{row_idx}"
        )
        ws.cell(row=row_idx, column=col_tot_paga, value=formula)

    # Formati
    for col_idx, header in enumerate(headers, start=1):
        if isinstance(header, date):
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0.00"
            ws.column_dimensions[get_column_letter(col_idx)].width = 10
        elif header in {
            "NETTO ORARIO",
            "NETTO MESE",
            "MAGGIORAZIONE FESTIVO",
            "ASSENZE",
            "ARRETRATO",
            "EXTRA",
            "AFFIANCAMENTO",
            "DOMENICHE",
            "RIMBORSI",
            "TOT. MESE DA PAGARE",
        }:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0.00"
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def safe_filename(text: str) -> str:
    text = normalize_text(text)
    bad = '\\/:*?"<>|'
    for ch in bad:
        text = text.replace(ch, "_")
    return text.strip().replace("  ", " ")


def build_single_sheet_pdf_bytes(record: dict) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=18,
        rightMargin=18,
        topMargin=18,
        bottomMargin=18,
    )

    styles = getSampleStyleSheet()
    story = []

    titolo = f"{normalize_text(record.get('nome', ''))} - {normalize_text(record.get('societa', ''))} - {normalize_text(record.get('attivita_riga', ''))}"
    story.append(Paragraph(titolo, styles["Title"]))
    story.append(Spacer(1, 8))

    header_rows = [
        ["Società", normalize_text(record.get("societa", "")), "Attività", normalize_text(record.get("attivita_riga", ""))],
        ["Nome", normalize_text(record.get("nome", "")), "CF", normalize_text(record.get("cf", ""))],
        ["PDV", normalize_text(record.get("pdv", "")), "Netto ora", f"{safe_float(record.get('netto_ora', 0.0)):.2f}"],
        ["Tipo contratto", normalize_text(record.get("tipo_contratto", "")), "Scadenza", normalize_text(record.get("scadenza_contratto", ""))],
        ["Netto mese", f"{safe_float(record.get('netto_mese', 0.0)):.2f}", "Tot netto mese", f"{safe_float(record.get('tot_netto_mese', 0.0)):.2f}"],
    ]

    tbl_header = Table(header_rows, repeatRows=0)
    tbl_header.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.6, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(tbl_header)
    story.append(Spacer(1, 10))

    df = record["tabella"].copy()

    if normalize_upper(record.get("origine_master", "")) == ORIGINE_EDICOLA:
        body = [["Giorno", "Data", "Ore", "€", "Assenza", "Festivo", "Stato"]]
        for _, row in df.iterrows():
            body.append(
                [
                    int(row["GIORNO_NUM"]),
                    normalize_text(row["DATA"]),
                    f"{safe_float(row['EDICOLA_ORE']):.2f}",
                    f"{safe_float(row['EDICOLA_€']):.2f}",
                    normalize_text(row["EDICOLA_TIPO_ASSENZA"]),
                    "Sì" if bool(row["FESTIVO"]) else "No",
                    normalize_text(row["ROW_STATUS"]),
                ]
            )
    else:
        body = [[
            "Giorno", "Data",
            "Mond. Ore", "Mond. €", "Mond. Ass.",
            "Giunti Ore", "Giunti €", "Giunti Ass.",
            "Festivo", "Stato"
        ]]
        for _, row in df.iterrows():
            body.append(
                [
                    int(row["GIORNO_NUM"]),
                    normalize_text(row["DATA"]),
                    f"{safe_float(row['MONDADORI_ORE']):.2f}",
                    f"{safe_float(row['MONDADORI_€']):.2f}",
                    normalize_text(row["MONDADORI_TIPO_ASSENZA"]),
                    f"{safe_float(row['GIUNTI_ORE']):.2f}",
                    f"{safe_float(row['GIUNTI_€']):.2f}",
                    normalize_text(row["GIUNTI_TIPO_ASSENZA"]),
                    "Sì" if bool(row["FESTIVO"]) else "No",
                    normalize_text(row["ROW_STATUS"]),
                ]
            )

    tbl_body = Table(body, repeatRows=1)
    tbl_body.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EFEFEF")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(tbl_body)
    story.append(Spacer(1, 10))

    fondo_rows = [
        ["Arretrato", f"{safe_float(record.get('arretrati', 0.0)):.2f}"],
        ["Extra", f"{safe_float(record.get('extra', 0.0)):.2f}"],
        ["Affiancamento", f"{safe_float(record.get('affiancamenti', 0.0)):.2f}"],
        ["Domeniche", f"{safe_float(record.get('domeniche', 0.0)):.2f}"],
        ["Rimborsi", f"{safe_float(record.get('rimborso', 0.0)):.2f}"],
        ["Note del mese", normalize_text(record.get("note_generali", ""))],
    ]

    tbl_fondo = Table(fondo_rows, colWidths=[120, 500])
    tbl_fondo.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(tbl_fondo)

    doc.build(story)
    return buffer.getvalue()


def build_pdf_zip_bytes(records: list[dict]) -> bytes:
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for record in records:
            nome = safe_filename(record.get("nome", "Senza Nome"))
            societa = safe_filename(record.get("societa", ""))
            attivita = safe_filename(record.get("attivita_riga", ""))
            filename = f"{nome}_{societa}_{attivita}.pdf"

            pdf_bytes = build_single_sheet_pdf_bytes(record)
            zf.writestr(filename, pdf_bytes)

    return zip_buffer.getvalue()


def filter_records_for_pdf_export(
    anno: int,
    mese: int,
    societa: str,
    attivita: str,
    tipo_download: str,
    nominativo: str | None = None,
) -> list[dict]:
    records = []

    for foglio_key in get_month_sheet_keys(anno, mese):
        record = st.session_state["fogli_generati"][foglio_key]

        societa_record = normalize_text(record.get("societa", ""))
        origine_master = normalize_upper(record.get("origine_master", ""))

        if societa != "Tutte" and normalize_upper(societa) != normalize_upper(societa_record):
            continue

        if attivita != "Tutte":
            if attivita == "Edicola" and origine_master != ORIGINE_EDICOLA:
                continue
            if attivita == "Libri" and origine_master != ORIGINE_LIBRI:
                continue

        if tipo_download == "Dipendente":
            if normalize_upper(record.get("nome", "")) != normalize_upper(nominativo or ""):
                continue

        records.append(record)

    return records

# =========================
# PAGINE
# =========================

def render_master_page():
    render_page_title("1. Origine dati")
    st.markdown('<div class="section-box">', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    with col1:
        file_edicola = st.file_uploader("Master Edicola", type=["xlsx"], key="upload_edicola")
    with col2:
        file_libri = st.file_uploader("Master Libri", type=["xlsx"], key="upload_libri")
    with col3:
        file_spot = st.file_uploader("Master Sostituti Spot", type=["xlsx"], key="upload_spot")

    if st.button("Verifica e carica master", type="primary", use_container_width=True):
        if not file_edicola or not file_libri or not file_spot:
            st.error("Vanno caricati tutti e 3 i master reali.")
            st.stop()

        try:
            df_edicola_raw = pd.read_excel(file_edicola)
            df_libri_raw = pd.read_excel(file_libri)
            df_spot_raw = pd.read_excel(file_spot)
        except Exception as exc:
            st.error(f"Errore nella lettura dei file Excel: {exc}")
            st.stop()

        df_edicola_raw = clean_columns(df_edicola_raw)
        df_libri_raw = clean_columns(df_libri_raw)
        df_spot_raw = clean_columns(df_spot_raw)

        missing_edicola = validate_columns(df_edicola_raw, COLONNE_MASTER_EDICOLA)
        missing_libri = validate_columns(df_libri_raw, COLONNE_MASTER_LIBRI)
        missing_spot = validate_columns(df_spot_raw, COLONNE_MASTER_SPOT)

        if missing_edicola:
            st.error(f"Master Edicola non valido. Colonne mancanti: {', '.join(missing_edicola)}")
            st.stop()
        if missing_libri:
            st.error(f"Master Libri non valido. Colonne mancanti: {', '.join(missing_libri)}")
            st.stop()
        if missing_spot:
            st.error(f"Master Sostituti Spot non valido. Colonne mancanti: {', '.join(missing_spot)}")
            st.stop()

        st.session_state["df_edicola"] = normalize_master_edicola(df_edicola_raw)
        st.session_state["df_libri"] = normalize_master_libri(df_libri_raw)
        st.session_state["df_spot"] = normalize_master_spot(df_spot_raw)
        st.session_state["generation_table"] = build_generation_table(
            st.session_state["df_edicola"],
            st.session_state["df_libri"],
        )
        st.session_state["fogli_generati"] = {}
        st.session_state["master_loaded"] = True
        st.success("Master verificati e caricati correttamente.")
        st.rerun()

    if st.session_state["master_loaded"]:
        with st.expander("Visualizza intestazioni reali e anteprima master"):
            st.write("**Master Edicola**")
            st.write(list(st.session_state["df_edicola"].columns))
            st.dataframe(st.session_state["df_edicola"].head(10), use_container_width=True)

            st.write("**Master Libri**")
            st.write(list(st.session_state["df_libri"].columns))
            st.dataframe(st.session_state["df_libri"].head(10), use_container_width=True)

            st.write("**Master Sostituti Spot**")
            st.write(list(st.session_state["df_spot"].columns))
            st.dataframe(st.session_state["df_spot"].head(10), use_container_width=True)

    st.markdown("</div>", unsafe_allow_html=True)


def render_generation_page():
    render_page_title("2. Generazione fogli")
    st.markdown('<div class="section-box">', unsafe_allow_html=True)

    if "step4_table" not in st.session_state:
        st.session_state["step4_table"] = pd.DataFrame(
            columns=[
                "ROW_ID",
                "STATO_FOGLIO",
                "ORIGINE_MASTER",
                "TIPO_LIBRI",
                "NOME",
                "COD_FISCALE",
                "SOCIETA",
                "PDV",
                "TELEFONO",
                "EMAIL",
                "NETTO_ORA",
                "NETTO_MESE",
                "TIPO_CONTRATTO",
                "SCADENZA_CONTRATTO",
                "ATTIVITA_RIGA",
            ]
        )

    if not st.session_state["master_loaded"]:
        st.warning("Prima carica e verifica i master nella sezione 'Origine dati'.")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    anno_corrente, mese_corrente = get_mese_anno_lavoro()

    anni_disponibili = [2025, 2026, 2027]
    mesi_disponibili = list(MESI.keys())

    idx_anno = anni_disponibili.index(anno_corrente) if anno_corrente in anni_disponibili else 1
    idx_mese = mesi_disponibili.index(mese_corrente) if mese_corrente in mesi_disponibili else 0

    col1, col2 = st.columns(2)
    with col1:
        anno = st.selectbox("Anno", anni_disponibili, index=idx_anno, key="anno_step3")
    with col2:
        mese = st.selectbox("Mese", mesi_disponibili, format_func=lambda x: MESI[x], index=idx_mese, key="mese_step3")

    st.session_state["anno_lavoro"] = int(anno)
    st.session_state["mese_lavoro"] = int(mese)

    # =========================
    # TABELLA PRINCIPALE DA MASTER
    # =========================
    full_table = st.session_state["generation_table"].copy()
    if full_table.empty:
        st.warning("Nessuna riga disponibile.")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    def is_created(row_id: str) -> bool:
        key = format_sheet_key(normalize_text(row_id), anno, mese)
        return key in st.session_state["fogli_generati"]

    full_table["STATO_FOGLIO"] = full_table["ROW_ID"].apply(
        lambda rid: "CREATO" if is_created(rid) else "NON CREATO"
    )

    met1, met2, met3 = st.columns(3)
    with met1:
        st.metric("Totale righe mese", len(full_table))
    with met2:
        st.metric("Fogli creati", int((full_table["STATO_FOGLIO"] == "CREATO").sum()))
    with met3:
        st.metric("Fogli mancanti", int((full_table["STATO_FOGLIO"] == "NON CREATO").sum()))

    st.markdown('<div class="inner-box">', unsafe_allow_html=True)
    st.markdown('<div class="table-title">Tabella principale fogli del mese</div>', unsafe_allow_html=True)

    colf1, colf2 = st.columns(2)
    with colf1:
        filtro_nome = st.text_input(
            "Filtro semplice nome/cognome/PDV",
            placeholder="Nome o PDV",
            key="filtro_gen",
        )
    with colf2:
        filtro_stato = st.selectbox(
            "Filtro stato",
            ["TUTTI", "CREATO", "NON CREATO"],
            key="filtro_stato",
        )

    table = full_table.copy()

    if filtro_nome.strip():
        up = filtro_nome.strip().upper()
        mask = (
            table["NOME"].astype(str).str.upper().str.contains(up, na=False)
            | table["PDV"].astype(str).str.upper().str.contains(up, na=False)
        )
        table = table[mask].copy()

    if filtro_stato != "TUTTI":
        table = table[table["STATO_FOGLIO"] == filtro_stato].copy()

    if table.empty:
        st.info("Nessun risultato trovato con i filtri inseriti.")
    else:
        table = table.reset_index(drop=True)

        start_row_idx = st.selectbox(
            "Riga di partenza per selezione assistita (max 50)",
            options=list(table.index),
            format_func=lambda i: (
                f"{table.loc[i, 'NOME']} | {table.loc[i, 'PDV']} | {table.loc[i, 'ATTIVITA_RIGA']} | {table.loc[i, 'STATO_FOGLIO']}"
            ),
            key="start_row_idx",
        )

        csel1, csel2 = st.columns(2)
        with csel1:
            if st.button("Seleziona 50 righe da qui in giù", use_container_width=True, key="sel50"):
                row_ids = table.loc[start_row_idx:, "ROW_ID"].tolist()[:50]
                base = st.session_state["generation_table"].copy()
                base.loc[base["ROW_ID"].isin(row_ids), "SELEZIONA"] = True
                st.session_state["generation_table"] = base
                st.rerun()

        with csel2:
            if st.button("Deseleziona righe visibili", use_container_width=True, key="desel_vis"):
                row_ids = table["ROW_ID"].tolist()
                base = st.session_state["generation_table"].copy()
                base.loc[base["ROW_ID"].isin(row_ids), "SELEZIONA"] = False
                st.session_state["generation_table"] = base
                st.rerun()

        editor_table = st.session_state["generation_table"].copy()
        editor_table["STATO_FOGLIO"] = editor_table["ROW_ID"].apply(
            lambda rid: "CREATO" if is_created(rid) else "NON CREATO"
        )

        if filtro_nome.strip():
            up = filtro_nome.strip().upper()
            mask = (
                editor_table["NOME"].astype(str).str.upper().str.contains(up, na=False)
                | editor_table["PDV"].astype(str).str.upper().str.contains(up, na=False)
            )
            editor_table = editor_table[mask].copy()

        if filtro_stato != "TUTTI":
            editor_table = editor_table[editor_table["STATO_FOGLIO"] == filtro_stato].copy()

        edited = st.data_editor(
            editor_table.reset_index(drop=True),
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            column_config={
                "SELEZIONA": st.column_config.CheckboxColumn("Seleziona"),
                "ROW_ID": None,
                "MASTER_INDEX": None,
                "IS_STEP4": None,
                "ORIGINE": st.column_config.TextColumn("Origine", disabled=True),
                "ORIGINE_MASTER": st.column_config.TextColumn("Attività base", disabled=True),
                "TIPO_LIBRI": st.column_config.TextColumn("Tipo libri", disabled=True),
                "STATO_FOGLIO": st.column_config.TextColumn("Stato foglio", disabled=True),
                "NOME": st.column_config.TextColumn("Nome", disabled=True),
                "COD_FISCALE": st.column_config.TextColumn("CF", disabled=True),
                "SOCIETA": st.column_config.TextColumn("Società", disabled=True),
                "PDV": st.column_config.TextColumn("PDV", disabled=True),
                "TELEFONO": st.column_config.TextColumn("Telefono", disabled=True),
                "EMAIL": st.column_config.TextColumn("Email", disabled=True),
                "NETTO_ORA": st.column_config.NumberColumn("Netto orario", format="%.2f €", disabled=True),
                "NETTO_MESE": st.column_config.NumberColumn("Netto mese", format="%.2f €", disabled=True),
                "TIPO_CONTRATTO": st.column_config.TextColumn("Tipo contratto", disabled=True),
                "SCADENZA_CONTRATTO": st.column_config.TextColumn("Scadenza contratto", disabled=True),
                "ATTIVITA_RIGA": st.column_config.TextColumn("Attività", disabled=True),
            },
            key="editor_generation_table_finale",
        )

        updated = st.session_state["generation_table"].copy()
        for _, row in edited[["ROW_ID", "SELEZIONA"]].iterrows():
            updated.loc[updated["ROW_ID"] == row["ROW_ID"], "SELEZIONA"] = bool(row["SELEZIONA"])
        st.session_state["generation_table"] = updated

        if st.button("Genera fogli presenze selezionati", type="primary", use_container_width=True, key="genera_fogli"):
            selected = st.session_state["generation_table"][
                st.session_state["generation_table"]["SELEZIONA"] == True
            ].copy()

            if selected.empty:
                st.warning("Seleziona almeno una riga.")
                st.markdown("</div>", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)
                return

            if len(selected) > 50:
                st.error("Puoi generare al massimo 50 fogli per volta.")
                st.markdown("</div>", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)
                return

            last_key = None
            for _, row in selected.iterrows():
                key = format_sheet_key(normalize_text(row["ROW_ID"]), anno, mese)

                if key in st.session_state["fogli_generati"]:
                    continue

                if bool(row.get("IS_STEP4", False)):
                    continue

                record = init_sheet_record_from_master(
                    row=row,
                    df_edicola=st.session_state["df_edicola"],
                    df_libri=st.session_state["df_libri"],
                    anno=anno,
                    mese=mese,
                )
                st.session_state["fogli_generati"][key] = record
                st.session_state["sheet_warnings"][key] = []
                last_key = key

            if last_key:
                st.session_state["foglio_attivo"] = last_key

            base = st.session_state["generation_table"].copy()
            base.loc[base["ROW_ID"].isin(selected["ROW_ID"].tolist()), "SELEZIONA"] = False
            st.session_state["generation_table"] = base

            st.success("Generazione fogli completata correttamente.")
            st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

def render_step4_page():
    render_page_title("4. Sostituzioni/Azzeramenti")
    st.markdown('<div class="section-box">', unsafe_allow_html=True)

    if not st.session_state["master_loaded"]:
        st.warning("Prima carica e verifica i master nella sezione 'Origine dati'.")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    anno, mese = get_mese_anno_lavoro()
    societa_options = get_societa_options()

    # =========================
    # NUOVO FOGLIO PRESENZA / SOSTITUZIONE
    # =========================
    st.markdown('<div class="action-box">', unsafe_allow_html=True)
    st.markdown('<div class="action-title">NUOVO FOGLIO PRESENZA / SOSTITUZIONE</div>', unsafe_allow_html=True)
    st.markdown('<div class="action-subtitle">Crea un foglio nuovo fuori dal flusso standard</div>', unsafe_allow_html=True)

    if not societa_options:
        st.error("Nessuna società disponibile nei master caricati.")
    else:
        c1, c2 = st.columns(2)
        with c1:
            modalita = st.selectbox(
                "Cosa vuoi fare",
                [
                    "Sostituzione con titolare",
                    "Sostituzione con sostituto spot",
                    "Foglio presenza vuoto",
                ],
                key="step4_modalita",
            )
        with c2:
            societa = st.selectbox("Società", societa_options, key="step4_societa")

        c3, c4 = st.columns(2)
        with c3:
            attivita = st.selectbox("Attività", [ORIGINE_EDICOLA, ORIGINE_LIBRI], key="step4_attivita")
        with c4:
            tipo_giorno = st.selectbox(
                "Giorno singolo oppure periodo",
                ["Giorno singolo", "Periodo"],
                key="step4_tipo_giorno",
            )

        primo_giorno = date(anno, mese, 1)
        ultimo_giorno = date(anno, mese, calendar.monthrange(anno, mese)[1])

        # Range ampio per permettere navigazione calendario,
        # ma default allineato a mese/anno del sistema.
        min_data_cal = date(anno - 1, 1, 1)
        max_data_cal = date(anno + 1, 12, 31)

        giorno_singolo_value = None
        periodo_value = None

        if tipo_giorno == "Giorno singolo":
            giorno_singolo_value = st.date_input(
                "Giorno singolo",
                value=primo_giorno,
                min_value=min_data_cal,
                max_value=max_data_cal,
                key="step4_giorno_singolo",
            )
        else:
            periodo_value = st.date_input(
                "Periodo",
                value=(primo_giorno, primo_giorno),
                min_value=min_data_cal,
                max_value=max_data_cal,
                key="step4_periodo",
            )

        selected_days = get_selected_days(
            anno=anno,
            mese=mese,
            modo=tipo_giorno,
            giorno_singolo_value=giorno_singolo_value,
            periodo_value=periodo_value,
        )

        dipendenti_table = pd.DataFrame()
        pdv_table = pd.DataFrame()

        if normalize_upper(modalita) != "FOGLIO PRESENZA VUOTO":
            dipendenti_table = build_dipendenti_table(societa, attivita, modalita)
            pdv_table = build_pdv_table(societa, attivita)

            filtro_dip = st.text_input(
                "Filtro tabella dipendenti",
                placeholder="Nome / cognome",
                key="step4_filtro_dip",
            )
            filtro_pdv = st.text_input(
                "Filtro tabella punto vendita",
                placeholder="PDV",
                key="step4_filtro_pdv",
            )

            if filtro_dip.strip() and not dipendenti_table.empty:
                mask = dipendenti_table["NOMINATIVO"].astype(str).str.upper().str.contains(
                    filtro_dip.strip().upper(),
                    na=False,
                )
                dipendenti_table = dipendenti_table[mask].copy()

            if filtro_pdv.strip() and not pdv_table.empty:
                mask = pdv_table["PUNTO_VENDITA"].astype(str).str.upper().str.contains(
                    filtro_pdv.strip().upper(),
                    na=False,
                )
                pdv_table = pdv_table[mask].copy()

            if not dipendenti_table.empty:
                dipendenti_table = dipendenti_table.sort_values(
                    ["NOMINATIVO", "CODICE_FISCALE"]
                ).reset_index(drop=True)

            if not pdv_table.empty:
                pdv_table = pdv_table.sort_values(
                    ["PUNTO_VENDITA", "TIPO_ATTIVITA"]
                ).reset_index(drop=True)

            st.markdown(
                """
                <div class="soft-note">
                    Selezione guidata:
                    <br>- puoi selezionare una sola riga per tabella
                    <br>- puoi selezionare entrambe le tabelle oppure una sola
                    <br>- se selezioni solo Dipendente, il PDV resta vuoto e le ORE saranno manuali
                    <br>- se selezioni solo Punto Vendita, il NOME può restare vuoto ma le ORE devono essere compilate dal sistema
                </div>
                """,
                unsafe_allow_html=True,
            )

            st.markdown('<div class="table-title">Tabella dipendenti</div>', unsafe_allow_html=True)
            if dipendenti_table.empty:
                st.info("Nessun dato disponibile nella tabella dipendenti.")
            else:
                dipendenti_table = st.data_editor(
                    dipendenti_table,
                    hide_index=True,
                    use_container_width=True,
                    num_rows="fixed",
                    column_config={
                        "SELEZIONA": st.column_config.CheckboxColumn("Seleziona"),
                        "SOURCE_ID": None,
                        "TEL": None,
                        "MAIL": None,
                        "NOMINATIVO": st.column_config.TextColumn("Nominativo", disabled=True),
                        "CODICE_FISCALE": st.column_config.TextColumn("Codice Fiscale", disabled=True),
                        "TIPO_CONTRATTO": st.column_config.TextColumn("Tipo contratto", disabled=True),
                        "SCADENZA_CONTRATTO": st.column_config.TextColumn("Scadenza contratto", disabled=True),
                        "NETTO_ORA": st.column_config.NumberColumn("Netto orario", format="%.2f", disabled=True),
                    },
                    key="step4_table_dipendenti",
                )

            st.markdown('<div class="table-title">Tabella punto vendita</div>', unsafe_allow_html=True)
            if pdv_table.empty:
                st.info("Nessun dato disponibile nella tabella punto vendita.")
            else:
                pdv_table = st.data_editor(
                    pdv_table,
                    hide_index=True,
                    use_container_width=True,
                    num_rows="fixed",
                    column_config={
                        "SELEZIONA": st.column_config.CheckboxColumn("Seleziona"),
                        "SOURCE_ID": None,
                        "MASTER_INDEX": None,
                        "TIPO_LIBRI": None,
                        "TIPO_ATTIVITA": st.column_config.TextColumn("Tipo attività", disabled=True),
                        "PUNTO_VENDITA": st.column_config.TextColumn("Punto vendita", disabled=True),
                        "LUNEDI": st.column_config.NumberColumn("Lunedì", format="%.2f", disabled=True),
                        "MARTEDI": st.column_config.NumberColumn("Martedì", format="%.2f", disabled=True),
                        "MERCOLEDI": st.column_config.NumberColumn("Mercoledì", format="%.2f", disabled=True),
                        "GIOVEDI": st.column_config.NumberColumn("Giovedì", format="%.2f", disabled=True),
                        "VENERDI": st.column_config.NumberColumn("Venerdì", format="%.2f", disabled=True),
                        "SABATO": st.column_config.NumberColumn("Sabato", format="%.2f", disabled=True),
                        "DOMENICA": st.column_config.NumberColumn("Domenica", format="%.2f", disabled=True),
                    },
                    key="step4_table_pdv",
                )

        if st.button(
            "NUOVO FOGLIO PRESENZA / SOSTITUZIONE",
            type="primary",
            use_container_width=True,
            key="btn_step4_generate",
        ):
            if not selected_days:
                st.error("Giorno o periodo non valido per il mese selezionato.")
                st.markdown("</div>", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)
                return

            record = base_step4_record(modalita, societa, attivita, anno, mese, selected_days)

            if normalize_upper(modalita) == "FOGLIO PRESENZA VUOTO":
                # foglio completamente manuale
                record["allow_increase"] = True
                record["netto_mese_dynamic"] = True
            else:
                dip_row, dip_count = get_single_selected_row(
                    dipendenti_table if not dipendenti_table.empty else pd.DataFrame()
                )
                pdv_row, pdv_count = get_single_selected_row(
                    pdv_table if not pdv_table.empty else pd.DataFrame()
                )

                if dip_count > 1 or pdv_count > 1:
                    st.error("Selezionare una sola riga per ogni tabella.")
                    st.markdown("</div>", unsafe_allow_html=True)
                    st.markdown("</div>", unsafe_allow_html=True)
                    return

                if dip_count == 0 and pdv_count == 0:
                    st.error("Selezionare almeno una riga su una delle due tabelle.")
                    st.markdown("</div>", unsafe_allow_html=True)
                    st.markdown("</div>", unsafe_allow_html=True)
                    return

                apply_selected_rows_to_record(record, dip_row, pdv_row, selected_days)

                # I fogli STEP 4 devono sempre restare compilabili manualmente nel corpo
                record["allow_increase"] = True
                record["netto_mese_dynamic"] = False

            update_sheet_totals(record)

            key = format_sheet_key(record["row_id"], anno, mese)
            st.session_state["fogli_generati"][key] = record
            st.session_state["sheet_warnings"][key] = []
            st.session_state["foglio_attivo"] = key
            append_record_to_generation_table(record)

            st.success("Foglio creato correttamente.")
            st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)

    # =========================
    # AZZERAMENTO MASSIVO
    # =========================
    st.markdown('<div class="action-box">', unsafe_allow_html=True)
    st.markdown('<div class="action-title">AZZERAMENTO MASSIVO</div>', unsafe_allow_html=True)
    st.markdown('<div class="action-subtitle">Azzera le ore di più fogli insieme senza aprirli uno a uno</div>', unsafe_allow_html=True)

    if not societa_options:
        st.info("Nessuna società disponibile.")
        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)
        return

    m1, m2 = st.columns(2)
    with m1:
        step_societa = st.selectbox("Società", societa_options, key="step525_societa")
    with m2:
        step_attivita = st.selectbox("Attività", [ORIGINE_EDICOLA, ORIGINE_LIBRI], key="step525_attivita")

    m3, m4 = st.columns(2)
    with m3:
        step_tipo_giorno = st.selectbox(
            "Giorno singolo oppure periodo",
            ["Giorno singolo", "Periodo"],
            key="step525_tipo_giorno",
        )
    with m4:
        primo_giorno_massive = date(anno, mese, 1)
        ultimo_giorno_massive = date(anno, mese, calendar.monthrange(anno, mese)[1])

        # Range ampio per permettere navigazione calendario,
        # ma default allineato a mese/anno del sistema.
        min_data_cal_massive = date(anno - 1, 1, 1)
        max_data_cal_massive = date(anno + 1, 12, 31)

        giorno_singolo_value_massive = None
        periodo_value_massive = None

        if step_tipo_giorno == "Giorno singolo":
            giorno_singolo_value_massive = st.date_input(
                "Giorno singolo",
                value=primo_giorno_massive,
                min_value=min_data_cal_massive,
                max_value=max_data_cal_massive,
                key="step525_giorno_singolo",
            )
        else:
            periodo_value_massive = st.date_input(
                "Periodo",
                value=(primo_giorno_massive, primo_giorno_massive),
                min_value=min_data_cal_massive,
                max_value=max_data_cal_massive,
                key="step525_periodo",
            )

    selected_days_massive = get_selected_days(
        anno=anno,
        mese=mese,
        modo=step_tipo_giorno,
        giorno_singolo_value=giorno_singolo_value_massive,
        periodo_value=periodo_value_massive,
    )

    righe_massive = []
    for foglio_key, record in st.session_state["fogli_generati"].items():
        if int(record.get("anno", 0)) != int(anno) or int(record.get("mese", 0)) != int(mese):
            continue
        if normalize_upper(record.get("societa", "")) != normalize_upper(step_societa):
            continue
        if normalize_upper(record.get("origine_master", "")) != normalize_upper(step_attivita):
            continue

        righe_massive.append(
            {
                "SELEZIONA": False,
                "FOGLIO_KEY": foglio_key,
                "NOMINATIVO_DIPENDENTE": normalize_text(record.get("nome", "")),
                "PDV": normalize_text(record.get("pdv", "")),
                "SOCIETA": normalize_text(record.get("societa", "")),
                "TIPO_ATTIVITA": normalize_text(record.get("attivita_riga", "")),
                "ORIGINE": "Nuovo foglio/sostituzione" if bool(record.get("is_step4", False)) else "Master",
            }
        )

    massive_table = pd.DataFrame(righe_massive)

    st.markdown(
        '<div class="soft-note"><b>Seleziona i fogli presenza da azzerare</b></div>',
        unsafe_allow_html=True,
    )

    if massive_table.empty:
        st.info("Nessun foglio presenza coerente trovato per i criteri selezionati.")
    else:
        massive_table = massive_table.sort_values(
            ["NOMINATIVO_DIPENDENTE", "PDV", "TIPO_ATTIVITA", "ORIGINE"]
        ).reset_index(drop=True)

        edited_massive = st.data_editor(
            massive_table,
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            column_config={
                "SELEZIONA": st.column_config.CheckboxColumn("Seleziona"),
                "FOGLIO_KEY": None,
                "NOMINATIVO_DIPENDENTE": st.column_config.TextColumn("Nominativo dipendente", disabled=True),
                "PDV": st.column_config.TextColumn("PDV", disabled=True),
                "SOCIETA": st.column_config.TextColumn("Società", disabled=True),
                "TIPO_ATTIVITA": st.column_config.TextColumn("Tipo attività", disabled=True),
                "ORIGINE": st.column_config.TextColumn("Origine", disabled=True),
            },
            key="step525_massive_table",
        )

        selected_rows = edited_massive[edited_massive["SELEZIONA"] == True].copy()
        count_selected = len(selected_rows)

        if selected_days_massive and count_selected > 0:
            st.markdown(
                f"""
                <div class="soft-note">
                    <b>Conferma operativa</b><br>
                    Fogli selezionati: {count_selected}<br>
                    Intervallo da azzerare: {selected_days_text(selected_days_massive)}
                </div>
                """,
                unsafe_allow_html=True,
            )

        if st.button("AZZERA ORE", type="primary", use_container_width=True, key="step525_btn_azzera"):
            if not selected_days_massive:
                st.error("Giorno o periodo non valido per il mese selezionato.")
                st.markdown("</div>", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)
                return

            if count_selected == 0:
                st.error("Seleziona almeno un foglio presenza dalla tabella.")
                st.markdown("</div>", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)
                return

            modificati = 0

            for _, row in selected_rows.iterrows():
                foglio_key = row["FOGLIO_KEY"]

                if foglio_key not in st.session_state["fogli_generati"]:
                    continue

                record = st.session_state["fogli_generati"][foglio_key]

                if record.get("lucchetto_mese", False) or record.get("lucchetto_foglio", False):
                    continue

                df = record["tabella"].copy()

                for idx in df.index:
                    if int(df.at[idx, "GIORNO_NUM"]) not in selected_days_massive:
                        continue

                    if normalize_upper(record.get("origine_master", "")) == ORIGINE_EDICOLA:
                        df.at[idx, "EDICOLA_ORE"] = 0.0
                        df.at[idx, "EDICOLA_€"] = 0.0
                        df.at[idx, "EDICOLA_TIPO_ASSENZA"] = ""
                    else:
                        df.at[idx, "MONDADORI_ORE"] = 0.0
                        df.at[idx, "MONDADORI_€"] = 0.0
                        df.at[idx, "MONDADORI_TIPO_ASSENZA"] = ""
                        df.at[idx, "GIUNTI_ORE"] = 0.0
                        df.at[idx, "GIUNTI_€"] = 0.0
                        df.at[idx, "GIUNTI_TIPO_ASSENZA"] = ""

                record["tabella"] = df
                st.session_state["sheet_warnings"][foglio_key] = update_sheet_totals(record)
                modificati += 1

            st.success(f"Azzeramento massivo completato correttamente. Fogli modificati: {modificati}")
            st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

def render_sheet_page():
    render_page_title("3. Fogli presenza")
    st.markdown('<div class="section-box">', unsafe_allow_html=True)

    if not st.session_state["fogli_generati"]:
        st.warning("Prima genera almeno un foglio.")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    keys = list(st.session_state["fogli_generati"].keys())
    if st.session_state["foglio_attivo"] not in keys:
        st.session_state["foglio_attivo"] = keys[0]

    filtro = st.text_input(
        "Filtro semplice fogli (nome o PDV)",
        placeholder="Nome o PDV",
        key="filtro_fogli_presenza",
    )

    filtered_keys = keys
    if filtro.strip():
        up = filtro.strip().upper()
        filtered_keys = [
            k for k in keys
            if up in normalize_text(st.session_state["fogli_generati"][k]["nome"]).upper()
            or up in normalize_text(st.session_state["fogli_generati"][k]["pdv"]).upper()
        ]

    if not filtered_keys:
        st.info("Nessun foglio trovato con il filtro inserito.")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    if st.session_state["foglio_attivo"] not in filtered_keys:
        st.session_state["foglio_attivo"] = filtered_keys[0]

    selected_key = st.selectbox(
        "Seleziona foglio attivo",
        options=filtered_keys,
        index=filtered_keys.index(st.session_state["foglio_attivo"]),
        format_func=lambda k: (
            f"{st.session_state['fogli_generati'][k]['nome'] or 'SENZA NOME'} | "
            f"{st.session_state['fogli_generati'][k]['pdv'] or 'SENZA PDV'} | "
            f"{st.session_state['fogli_generati'][k]['attivita_riga']} | "
            f"{MESI[st.session_state['fogli_generati'][k]['mese']]} {st.session_state['fogli_generati'][k]['anno']}"
        ),
        key="select_foglio_attivo_unificato",
    )

    st.session_state["foglio_attivo"] = selected_key
    record = st.session_state["fogli_generati"][selected_key]
    locked = record["lucchetto_mese"] or record["lucchetto_foglio"]
    is_step4 = bool(record.get("is_step4", False))

    warnings = st.session_state["sheet_warnings"].get(selected_key, [])
    if warnings:
        for msg in warnings:
            render_warning_box(msg)

    # =========================
    # INTESTAZIONE
    # =========================
    st.markdown('<div class="inner-box">', unsafe_allow_html=True)

    if is_step4:
        st.markdown(
            """
            <div class="soft-note">
                Foglio creato con Nuovo foglio / Sostituzione.
                Intestazione semplificata attiva.
            </div>
            """,
            unsafe_allow_html=True,
        )

        col_h1, col_h2, col_h3 = st.columns(3)
        with col_h1:
            record["nome"] = st.text_input(
                "Nome",
                value=record["nome"],
                disabled=locked,
                key=f"step4_nome_{selected_key}",
            )
            record["cf"] = st.text_input(
                "Codice Fiscale",
                value=record["cf"],
                disabled=locked,
                key=f"step4_cf_{selected_key}",
            )
        with col_h2:
            record["pdv"] = st.text_input(
                "PDV",
                value=record["pdv"],
                disabled=locked,
                key=f"step4_pdv_{selected_key}",
            )
            record["tipo_contratto"] = st.text_input(
                "Tipo contratto",
                value=record["tipo_contratto"],
                disabled=locked,
                key=f"step4_tipo_contratto_{selected_key}",
            )
        with col_h3:
            record["scadenza_contratto"] = st.text_input(
                "Scadenza contratto",
                value=record["scadenza_contratto"],
                disabled=locked,
                key=f"step4_scadenza_{selected_key}",
            )
            record["netto_ora"] = st.number_input(
                "Netto ora",
                value=float(record["netto_ora"]),
                step=0.10,
                disabled=locked,
                key=f"step4_netto_ora_{selected_key}",
            )

        col_lock1, col_lock2 = st.columns(2)
        with col_lock1:
            record["lucchetto_foglio"] = st.checkbox(
                "Lucchetto foglio",
                value=record["lucchetto_foglio"],
                disabled=record["lucchetto_mese"],
                key=f"step4_lock_foglio_{selected_key}",
            )
        with col_lock2:
            record["lucchetto_mese"] = st.checkbox(
                "Lucchetto mese",
                value=record["lucchetto_mese"],
                key=f"step4_lock_mese_{selected_key}",
            )

    else:
        col_h1, col_h2, col_h3, col_h4 = st.columns(4)
        with col_h1:
            st.text_input("Società", value=record["societa"], disabled=True, key=f"societa_{selected_key}")
            st.text_input("Attività", value=record["attivita_riga"], disabled=True, key=f"attivita_{selected_key}")
            st.text_input("Nome", value=record["nome"], disabled=True, key=f"nome_{selected_key}")
        with col_h2:
            st.text_input("CF", value=record["cf"], disabled=True, key=f"cf_{selected_key}")
            st.text_input("PDV", value=record["pdv"], disabled=True, key=f"pdv_{selected_key}")
            st.text_input("Mese", value=f"{MESI[record['mese']]} {record['anno']}", disabled=True, key=f"mese_{selected_key}")
        with col_h3:
            record["tipo_contratto"] = st.text_input(
                "Tipo contratto",
                value=record["tipo_contratto"],
                disabled=locked,
                key=f"tipo_contratto_{selected_key}",
            )
            record["scadenza_contratto"] = st.text_input(
                "Scadenza contratto",
                value=record["scadenza_contratto"],
                disabled=locked,
                key=f"scadenza_{selected_key}",
            )
            st.number_input("NETTO MESE", value=float(record["netto_mese"]), disabled=True, key=f"netto_mese_{selected_key}")
        with col_h4:
            record["netto_ora"] = st.number_input(
                "Netto orario",
                value=float(record["netto_ora"]),
                step=0.10,
                disabled=locked,
                key=f"netto_ora_{selected_key}",
            )
            st.number_input("Giorni lavorati", value=int(record["giorni_lavorati"]), disabled=True, key=f"giorni_lavorati_{selected_key}")
            st.number_input("Giorni modificati", value=int(record["giorni_modificati"]), disabled=True, key=f"giorni_modificati_{selected_key}")

        col_h5, col_h6, col_h7 = st.columns(3)
        with col_h5:
            st.number_input("Tot ore lavorative mese", value=float(record["tot_ore_lavorative_mese"]), disabled=True, key=f"tot_ore_{selected_key}")
        with col_h6:
            st.number_input("Tot ore ridotte/azzerate", value=float(record["tot_ore_azzerate"]), disabled=True, key=f"tot_ore_azzerate_{selected_key}")
        with col_h7:
            st.number_input("Tot € da scalare", value=float(record["tot_euro_da_scalare"]), disabled=True, key=f"tot_scalare_{selected_key}")

        col_lock1, col_lock2 = st.columns(2)
        with col_lock1:
            record["lucchetto_foglio"] = st.checkbox(
                "Lucchetto foglio",
                value=record["lucchetto_foglio"],
                disabled=record["lucchetto_mese"],
                key=f"lock_foglio_{selected_key}",
            )
        with col_lock2:
            record["lucchetto_mese"] = st.checkbox(
                "Lucchetto mese",
                value=record["lucchetto_mese"],
                key=f"lock_mese_{selected_key}",
            )

    st.markdown("</div>", unsafe_allow_html=True)

    locked = record["lucchetto_mese"] or record["lucchetto_foglio"]

    # =========================
    # CORPO FOGLIO
    # =========================
    st.markdown('<div class="inner-box">', unsafe_allow_html=True)
    st.markdown('<div class="table-title">Corpo foglio presenze</div>', unsafe_allow_html=True)

    origine = normalize_upper(record["origine_master"])
    prev_snapshot = record["tabella"].copy()

    if origine == ORIGINE_EDICOLA:
        visible_cols = [
            "GIORNO_NUM",
            "DATA_VIEW",
            "GIORNO_SETTIMANA",
            "EDICOLA_ORE",
            "EDICOLA_€",
            "EDICOLA_TIPO_ASSENZA",
            "FESTIVO",
        ]

        display_df = record["tabella"][COLONNE_TABELLA_TECNICHE].copy().reset_index(drop=True)

        edited = st.data_editor(
            display_df[visible_cols],
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            disabled=locked,
            height=1140,
            column_config={
                "GIORNO_NUM": st.column_config.NumberColumn("Giorno", disabled=True),
                "DATA_VIEW": st.column_config.TextColumn("Data", disabled=True),
                "GIORNO_SETTIMANA": st.column_config.TextColumn("Giorno settimana", disabled=True),
                "EDICOLA_ORE": st.column_config.NumberColumn("Edicola Ore", min_value=0.0, step=0.5, format="%.2f"),
                "EDICOLA_€": st.column_config.NumberColumn("Edicola €", format="%.2f €", disabled=True),
                "EDICOLA_TIPO_ASSENZA": st.column_config.SelectboxColumn("Tipo assenza", options=TIPI_ASSENZA),
                "FESTIVO": st.column_config.CheckboxColumn("Festivo"),
            },
            key=f"editor_tabella_{selected_key}",
        )

        if not locked:
            merged = prev_snapshot.copy().reset_index(drop=True)
            merged["EDICOLA_ORE"] = edited["EDICOLA_ORE"].apply(safe_float)
            merged["EDICOLA_TIPO_ASSENZA"] = edited["EDICOLA_TIPO_ASSENZA"].apply(normalize_text)
            merged["FESTIVO"] = edited["FESTIVO"].astype(bool)

            record["tabella"] = merged
            st.session_state["sheet_warnings"][selected_key] = update_sheet_totals(record)

            if not record["tabella"][visible_cols].reset_index(drop=True).equals(prev_snapshot[visible_cols].reset_index(drop=True)):
                st.rerun()

    else:
        visible_cols = [
            "GIORNO_NUM",
            "DATA_VIEW",
            "GIORNO_SETTIMANA",
            "MONDADORI_ORE",
            "MONDADORI_€",
            "MONDADORI_TIPO_ASSENZA",
            "GIUNTI_ORE",
            "GIUNTI_€",
            "GIUNTI_TIPO_ASSENZA",
            "FESTIVO",
        ]

        display_df = record["tabella"][COLONNE_TABELLA_TECNICHE].copy().reset_index(drop=True)

        edited = st.data_editor(
            display_df[visible_cols],
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            disabled=locked,
            height=1140,
            column_config={
                "GIORNO_NUM": st.column_config.NumberColumn("Giorno", disabled=True),
                "DATA_VIEW": st.column_config.TextColumn("Data", disabled=True),
                "GIORNO_SETTIMANA": st.column_config.TextColumn("Giorno settimana", disabled=True),
                "MONDADORI_ORE": st.column_config.NumberColumn("Mondadori Ore", min_value=0.0, step=0.5, format="%.2f"),
                "MONDADORI_€": st.column_config.NumberColumn("Mondadori €", format="%.2f €", disabled=True),
                "MONDADORI_TIPO_ASSENZA": st.column_config.SelectboxColumn("Assenza Mondadori", options=TIPI_ASSENZA),
                "GIUNTI_ORE": st.column_config.NumberColumn("Giunti Ore", min_value=0.0, step=0.5, format="%.2f"),
                "GIUNTI_€": st.column_config.NumberColumn("Giunti €", format="%.2f €", disabled=True),
                "GIUNTI_TIPO_ASSENZA": st.column_config.SelectboxColumn("Assenza Giunti", options=TIPI_ASSENZA),
                "FESTIVO": st.column_config.CheckboxColumn("Festivo"),
            },
            key=f"editor_tabella_{selected_key}",
        )

        if not locked:
            merged = prev_snapshot.copy().reset_index(drop=True)
            merged["MONDADORI_ORE"] = edited["MONDADORI_ORE"].apply(safe_float)
            merged["MONDADORI_TIPO_ASSENZA"] = edited["MONDADORI_TIPO_ASSENZA"].apply(normalize_text)
            merged["GIUNTI_ORE"] = edited["GIUNTI_ORE"].apply(safe_float)
            merged["GIUNTI_TIPO_ASSENZA"] = edited["GIUNTI_TIPO_ASSENZA"].apply(normalize_text)
            merged["FESTIVO"] = edited["FESTIVO"].astype(bool)

            record["tabella"] = merged
            st.session_state["sheet_warnings"][selected_key] = update_sheet_totals(record)

            if not record["tabella"][visible_cols].reset_index(drop=True).equals(prev_snapshot[visible_cols].reset_index(drop=True)):
                st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)

    # =========================
    # FONDO FOGLIO
    # =========================
    st.markdown('<div class="inner-box">', unsafe_allow_html=True)
    st.markdown('<div class="table-title">Fondo foglio</div>', unsafe_allow_html=True)

    prev_fondo = (
        float(record["arretrati"]),
        float(record["extra"]),
        float(record["affiancamenti"]),
        float(record["domeniche"]),
        float(record["rimborso"]),
        record["note_generali"],
    )

    col_b1, col_b2, col_b3, col_b4, col_b5 = st.columns(5)
    with col_b1:
        record["arretrati"] = st.number_input("€ arretrato", value=float(record["arretrati"]), step=0.50, disabled=locked, key=f"arretrati_{selected_key}")
    with col_b2:
        record["extra"] = st.number_input("€ extra", value=float(record["extra"]), step=0.50, disabled=locked, key=f"extra_{selected_key}")
    with col_b3:
        record["affiancamenti"] = st.number_input("€ affiancamento", value=float(record["affiancamenti"]), step=0.50, disabled=locked, key=f"affiancamenti_{selected_key}")
    with col_b4:
        record["domeniche"] = st.number_input("€ domeniche", value=float(record["domeniche"]), step=0.50, disabled=locked, key=f"domeniche_{selected_key}")
    with col_b5:
        record["rimborso"] = st.number_input("€ rimborso", value=float(record["rimborso"]), step=0.50, disabled=locked, key=f"rimborso_{selected_key}")

    uploaded_docs = st.file_uploader(
        "Allegati rimborso",
        accept_multiple_files=True,
        disabled=locked,
        key=f"rimborso_upload_{selected_key}",
    )
    if uploaded_docs is not None:
        record["rimborso_allegati"] = [file.name for file in uploaded_docs]

    if record["rimborso_allegati"]:
        st.caption("Allegati caricati: " + ", ".join(record["rimborso_allegati"]))

    record["note_generali"] = st.text_area(
        "NOTE GENERALI DEL MESE",
        value=record["note_generali"],
        height=130,
        disabled=locked,
        key=f"note_generali_{selected_key}",
    )

    update_sheet_totals(record)
    new_fondo = (
        float(record["arretrati"]),
        float(record["extra"]),
        float(record["affiancamenti"]),
        float(record["domeniche"]),
        float(record["rimborso"]),
        record["note_generali"],
    )
    if new_fondo != prev_fondo:
        st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)

    # =========================
    # TOTALE FINALE
    # =========================
    st.markdown('<div class="inner-box">', unsafe_allow_html=True)
    st.markdown('<div class="table-title">Totale finale</div>', unsafe_allow_html=True)

    col_t1, col_t2 = st.columns([2, 1])
    with col_t1:
        st.markdown(
            """
            <div class="soft-note">
                <b>Formula finale</b><br>
                TOT NETTO MESE = Totale attività corpo foglio + Arretrati + Extra + Affiancamenti + Domeniche + Rimborsi
            </div>
            """,
            unsafe_allow_html=True,
        )
    with col_t2:
        st.metric("TOT NETTO MESE", f"€ {record['tot_netto_mese']:.2f}")

    st.markdown("</div>", unsafe_allow_html=True)

    # =========================
    # TASTI
    # =========================
    col_reset1, col_reset2 = st.columns(2)
    with col_reset1:
        if st.button("Azzera ore foglio", disabled=locked, use_container_width=True, key=f"azzera_{selected_key}"):
            clear_entire_sheet(record)
            st.session_state["sheet_warnings"][selected_key] = []
            st.rerun()
    with col_reset2:
        if st.button("Elimina foglio presenza", disabled=locked, use_container_width=True, key=f"elimina_{selected_key}"):
            row_id = record["row_id"]
            del st.session_state["fogli_generati"][selected_key]
            st.session_state["sheet_warnings"].pop(selected_key, None)
            base = st.session_state["generation_table"].copy()
            base = base[base["ROW_ID"] != row_id].copy().reset_index(drop=True)
            st.session_state["generation_table"] = base

            remaining = list(st.session_state["fogli_generati"].keys())
            st.session_state["foglio_attivo"] = remaining[0] if remaining else None
            st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)

def render_chiusura_mese_page():
    render_page_title("5. Chiusura mese")
    st.markdown('<div class="section-box">', unsafe_allow_html=True)

    anno, mese = get_mese_anno_lavoro()
    table_df = build_chiusura_mese_table_df(anno, mese)

    if "contatore_estrazioni_singole" not in st.session_state:
        st.session_state["contatore_estrazioni_singole"] = 0
    if "contatore_estrazioni_massive" not in st.session_state:
        st.session_state["contatore_estrazioni_massive"] = 0

    # =========================
    # TABELLA DINAMICA DEL MESE
    # =========================
    st.markdown('<div class="inner-box">', unsafe_allow_html=True)
    st.markdown('<div class="table-title">Tabella dinamica fogli presenza del mese</div>', unsafe_allow_html=True)

    if table_df.empty:
        st.info("Nessun foglio presenza presente per il mese selezionato.")
    else:
        st.dataframe(table_df, use_container_width=True, hide_index=True)

    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    # =========================
    # CONTATORI
    # =========================
    totale_fogli_mese = len(get_month_sheet_keys(anno, mese))

    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Totale fogli mese", totale_fogli_mese)
    with c2:
        st.metric("Totale estrazioni singole", int(st.session_state["contatore_estrazioni_singole"]))
    with c3:
        st.metric("Totale estrazione massiva", int(st.session_state["contatore_estrazioni_massive"]))

    st.markdown("<br>", unsafe_allow_html=True)

    # =========================
    # DOWNLOAD FOGLI PDF
    # =========================
    st.markdown('<div class="action-box">', unsafe_allow_html=True)
    st.markdown('<div class="action-title">DOWNLOAD FOGLI PDF</div>', unsafe_allow_html=True)
    st.markdown('<div class="action-subtitle">Esporta i fogli presenza del mese in ZIP PDF</div>', unsafe_allow_html=True)

    d1, d2, d3 = st.columns(3)
    with d1:
        pdf_societa = st.selectbox("Società", ["Tutte", "Tekmar", "UP"], key="pdf_societa")
    with d2:
        pdf_attivita = st.selectbox("Tipo attività", ["Tutte", "Edicola", "Libri"], key="pdf_attivita")
    with d3:
        pdf_tipo_download = st.selectbox("Tipo download", ["Totale", "Dipendente"], key="pdf_tipo_download")

    nominativi_disponibili = sorted(
        list(
            {
                normalize_text(record.get("nome", ""))
                for foglio_key, record in st.session_state["fogli_generati"].items()
                if int(record.get("anno", 0)) == int(anno)
                and int(record.get("mese", 0)) == int(mese)
                and normalize_text(record.get("nome", "")) != ""
            }
        )
    )

    pdf_nominativo = None
    if pdf_tipo_download == "Dipendente":
        pdf_nominativo = st.selectbox(
            "Seleziona dipendente",
            nominativi_disponibili if nominativi_disponibili else [""],
            key="pdf_nominativo",
        )

    pdf_records = filter_records_for_pdf_export(
        anno=anno,
        mese=mese,
        societa=pdf_societa,
        attivita=pdf_attivita,
        tipo_download=pdf_tipo_download,
        nominativo=pdf_nominativo,
    )

    st.markdown(
        f"""
        <div class="soft-note">
            Fogli selezionati per il download PDF: <b>{len(pdf_records)}</b><br>
            Nota: lo ZIP dei giustificativi reali non è ancora disponibile perché oggi il sistema salva solo il nome allegato e non il file reale.
        </div>
        """,
        unsafe_allow_html=True,
    )

    if pdf_records:
        pdf_zip_bytes = build_pdf_zip_bytes(pdf_records)
        pdf_file_name = f"fogli_pdf_{anno}_{mese:02d}.zip"

        clicked_pdf = st.download_button(
            "DOWNLOAD FOGLI PDF",
            data=pdf_zip_bytes,
            file_name=pdf_file_name,
            mime="application/zip",
            use_container_width=True,
            key="btn_download_pdf_zip",
        )

        if clicked_pdf:
            if pdf_tipo_download == "Dipendente":
                st.session_state["contatore_estrazioni_singole"] += len(pdf_records)
            else:
                st.session_state["contatore_estrazioni_massive"] = len(pdf_records)
    else:
        st.info("Nessun foglio coerente trovato per il download PDF selezionato.")

    st.markdown("</div>", unsafe_allow_html=True)

    # =========================
    # DOWNLOAD RIEPILOGO MESE XLSX
    # =========================
    st.markdown('<div class="action-box">', unsafe_allow_html=True)
    st.markdown('<div class="action-title">DOWNLOAD RIEPILOGO MESE</div>', unsafe_allow_html=True)
    st.markdown('<div class="action-subtitle">Esporta il riepilogo totale del mese in formato Excel (.xlsx)</div>', unsafe_allow_html=True)

    export_rows = build_chiusura_mese_export_rows(anno, mese)

    st.markdown(
        f"""
        <div class="soft-note">
            Totale fogli inclusi nel riepilogo mese: <b>{len(export_rows)}</b>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if export_rows:
        xlsx_bytes = build_chiusura_mese_xlsx_bytes(anno, mese)
        clicked_xlsx = st.download_button(
            "DOWNLOAD RIEPILOGO MESE",
            data=xlsx_bytes,
            file_name=f"chiusura_mese_{anno}_{mese:02d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="btn_download_riepilogo_xlsx",
        )

        if clicked_xlsx:
            st.session_state["contatore_estrazioni_massive"] = len(export_rows)
    else:
        st.info("Nessun dato disponibile per il riepilogo mese.")

    st.markdown("</div>", unsafe_allow_html=True)

    # =========================
    # PULIZIA MESE
    # =========================
    st.markdown('<div class="action-box">', unsafe_allow_html=True)
    st.markdown('<div class="action-title">PULIZIA MESE</div>', unsafe_allow_html=True)
    st.markdown('<div class="action-subtitle">Elimina fogli del mese e master, ripulendo anche lo storage</div>', unsafe_allow_html=True)

    conferma_download = st.radio(
        "Hai effettuato i download del mese?",
        ["No", "Sì"],
        horizontal=True,
        key="pulizia_mese_conferma_download",
    )

    if st.button("PULIZIA MESE", type="primary", use_container_width=True, key="btn_pulizia_mese"):
        if conferma_download != "Sì":
            st.error("Prima effettua tutti i download per non perdere il lavoro del mese.")
        else:
            keys_to_delete = get_month_sheet_keys(anno, mese)
            for foglio_key in keys_to_delete:
                st.session_state["fogli_generati"].pop(foglio_key, None)
                st.session_state["sheet_warnings"].pop(foglio_key, None)

            st.session_state["df_edicola"] = pd.DataFrame()
            st.session_state["df_libri"] = pd.DataFrame()
            st.session_state["df_spot"] = pd.DataFrame()
            st.session_state["generation_table"] = pd.DataFrame()
            st.session_state["master_loaded"] = False
            st.session_state["foglio_attivo"] = None
            st.session_state["contatore_estrazioni_singole"] = 0
            st.session_state["contatore_estrazioni_massive"] = 0

            if STATE_FILE.exists():
                STATE_FILE.unlink()

            st.success("Pulizia mese completata correttamente.")
            st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

# =========================
# APP
# =========================

def main():
    ensure_session_state()
    inject_global_css()
    render_header()

    with st.sidebar:
        st.markdown('<div class="sidebar-logo-wrap">', unsafe_allow_html=True)
        st.image(LOGO_URL, width=170)
        st.markdown("</div>", unsafe_allow_html=True)

        sezione = st.radio(
            "Sezioni operative",
            [
                "Origine dati",
                "Generazione fogli",
                "Fogli presenza",
                "Sostituzioni/Azzeramenti",
                "Chiusura mese",
            ],
            key="main_navigation_radio",
        )

    if sezione == "Origine dati":
        render_master_page()
    elif sezione == "Generazione fogli":
        render_generation_page()
    elif sezione == "Fogli presenza":
        render_sheet_page()
    elif sezione == "Sostituzioni/Azzeramenti":
        render_step4_page()
    elif sezione == "Chiusura mese":
        render_chiusura_mese_page()

    save_app_state()

main()
